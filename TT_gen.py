import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from collections import defaultdict
import csv
import json
import os
import traceback
def load_config():
    """Load configuration from config.json"""
    try:
        with open('config.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print("Warning: config.json not found, using default values")
        return {
            "timetable_settings": {
                "working_days": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
                "start_time": "09:00",
                "end_time": "18:30",
                "slot_duration_minutes": 30
            },
            "break_settings": {
                "lunch_break": {
                    "enabled": True,
                    "start_time": "12:30",
                    "end_time": "14:00"
                }
            },
            "duration_constants": {
                "lecture_slots": 3,
                "lab_slots": 4,
                "tutorial_slots": 2
            }
        }

CONFIG = load_config()
def parse_time_config(time_str):
    """Convert time string from config (HH:MM) to time object"""
    hour, minute = map(int, time_str.split(':'))
    return time(hour, minute)
DAYS = CONFIG['timetable_settings']['working_days']
START_TIME = parse_time_config(CONFIG['timetable_settings']['start_time'])
END_TIME = parse_time_config(CONFIG['timetable_settings']['end_time'])
LECTURE_DURATION = CONFIG['duration_constants']['lecture_slots']
LAB_DURATION = CONFIG['duration_constants']['lab_slots']
TUTORIAL_DURATION = CONFIG['duration_constants']['tutorial_slots']
SELF_STUDY_DURATION = CONFIG['duration_constants'].get('self_study_slots', 2)
BREAK_DURATION = 1
lunch_config = CONFIG['break_settings']['lunch_break']
if lunch_config['enabled']:
    LUNCH_WINDOW_START = parse_time_config(lunch_config['start_time'])
    LUNCH_WINDOW_END = parse_time_config(lunch_config['end_time'])
    LUNCH_DURATION = lunch_config['duration_minutes']
else:
    LUNCH_WINDOW_START = None
    LUNCH_WINDOW_END = None
    LUNCH_DURATION = 0
elective_times = CONFIG['scheduling_preferences']['elective_slots']
ELECTIVE_SLOT_START_TIMES = [parse_time_config(t) for t in elective_times]
ELECTIVE_SLOT_INDICES = []
ELECTIVE_SLOT_SET = set()
ELECTIVE_USE_LAST = False
COLOR_PALETTE = [
    "FF5733", "33FF57", "3357FF", "FF33A8", "33FFF7", 
    "F7FF33", "FF33F7", "33F7FF", "FFB533", "B533FF",
    "33FFB5", "FF5F33", "335FFF", "B5FF33", "FF33B5"
]
try:
    rooms_df = pd.read_csv('rooms.csv')
    lecture_rooms = rooms_df[rooms_df['type'] == 'LECTURE_ROOM']['roomNumber'].tolist()
    computer_lab_rooms = rooms_df[rooms_df['type'] == 'COMPUTER_LAB']['roomNumber'].tolist()
    large_rooms = rooms_df[rooms_df['type'] == 'SEATER_120']['roomNumber'].tolist()
    if not lecture_rooms:
        print("Warning: No LECTURE_ROOM type rooms found in rooms.csv")
    if not computer_lab_rooms:
        print("Warning: No COMPUTER_LAB type rooms found in rooms.csv")
    if not large_rooms:
        print("Warning: No SEATER_120 type rooms found in rooms.csv")
except FileNotFoundError:
    print("Error: File 'rooms.csv' not found in the current directory")
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []
except Exception as e:
    print(f"Error loading rooms.csv: {e}")
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []

def generate_course_color():
    
    for color in COLOR_PALETTE:
        yield color
    while True:
        r = format(random.randint(180, 255), '02x')
        g = format(random.randint(180, 255), '02x')
        b = format(random.randint(180, 255), '02x')
        yield f"{r}{g}{b}"

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=30)
        slots.append((current, next_time.time()))
        current_time = next_time
    
    return slots
try:
    df = pd.read_csv('combined.csv')
except FileNotFoundError:
    print("Error: File 'combined.csv' not found in the current directory")
    exit()
def load_faculty_preferences():
    """Load faculty preferences from faculty_preferences.csv"""
    preferences = {}
    try:
        prefs_df = pd.read_csv('faculty_preferences.csv')
        for _, row in prefs_df.iterrows():
            faculty = str(row['Faculty'])
            course_code = str(row['Course Code'])
            day = str(row['Preferred Day'])
            time_str = str(row['Preferred Time'])
            priority = str(row['Priority']).upper()
            
            key = (faculty, course_code)
            if key not in preferences:
                preferences[key] = []
            
            preferences[key].append({
                'day': day,
                'time': time_str,
                'priority': priority
            })
        print(f"Loaded {len(preferences)} faculty preferences")
    except FileNotFoundError:
        print("Info: No faculty_preferences.csv found. Continuing without preferences.")
    except Exception as e:
        print(f"Warning: Error loading faculty preferences: {e}")
    
    return preferences
def load_reserved_slots():
    """Load reserved time slots from reserved_slots.json"""
    try:
        with open('reserved_slots.json', 'r') as f:
            data = json.load(f)
            reserved = {}
            for day_name, day_data in data.get('reserved_slots', {}).items():
                day_idx = DAYS.index(day_name) if day_name in DAYS else None
                if day_idx is not None:
                    reserved[day_idx] = {}
                    for dept, info in day_data.items():
                        semesters = info.get('semesters', [])
                        slots = info.get('slots', [])
                        reserved[day_idx][(dept, tuple(semesters))] = slots
            print(f"Loaded reserved slots for {len(reserved)} days")
            return reserved
    except FileNotFoundError:
        print("Info: No reserved_slots.json found. No slots reserved.")
        return {}
    except Exception as e:
        print(f"Warning: Error loading reserved slots: {e}")
        return {}
def load_teaching_assistants():
    """Load TA information from teaching_assistants.csv"""
    tas = []
    try:
        ta_df = pd.read_csv('teaching_assistants.csv')
        for _, row in ta_df.iterrows():
            tas.append({
                'name': str(row['TA Name']),
                'email': str(row['Email']),
                'department': str(row['Department']),
                'available_days': str(row['Available Days']).split(','),
                'available_times': str(row['Available Times']),
                'course_preferences': str(row['Course Preferences']).split(';'),
                'type': str(row['Type'])
            })
        print(f"Loaded {len(tas)} teaching assistants")
    except FileNotFoundError:
        print("Info: No teaching_assistants.csv found. No TAs will be allocated.")
    except Exception as e:
        print(f"Warning: Error loading teaching assistants: {e}")
    
    return tas

def is_break_time(slot):
    
    start, end = slot
    morning_break = False
    if CONFIG.get('break_settings', {}).get('morning_break', {}).get('enabled', False):
        morning_start = parse_time_config(CONFIG['break_settings']['morning_break']['start_time'])
        morning_duration = CONFIG['break_settings']['morning_break']['duration_minutes']
        morning_end_dt = datetime.combine(datetime.today(), morning_start) + timedelta(minutes=morning_duration)
        morning_end = morning_end_dt.time()
        morning_break = (morning_start <= start < morning_end)
    lunch_break = False
    if LUNCH_WINDOW_START and LUNCH_WINDOW_END:
        lunch_break = (LUNCH_WINDOW_START <= start < LUNCH_WINDOW_END)
    
    return morning_break or lunch_break

def check_professor_availability(professor_schedule, faculty, day, start_slot, duration, activity_type):
    
    existing_slots = sorted(list(professor_schedule[faculty][day]))
    if not existing_slots:
        return True  
    end_slot = start_slot + duration - 1
    MIN_GAP_SLOTS = 6  
    
    for slot in existing_slots:
        if start_slot <= slot <= end_slot:
            return False  
        if slot < start_slot and start_slot - slot < MIN_GAP_SLOTS:
            return False
            
        if slot > end_slot and slot - end_slot < MIN_GAP_SLOTS:
            return False
    
    return True

def check_professor_constraint(professor_schedule, faculty, day, start_slot, duration, timetable, time_slots):
    
    if not professor_schedule[faculty][day]:
        return True
    new_class_start_time = time_slots[start_slot][0]
    new_class_end_time = time_slots[start_slot + duration - 1][1]
    today = datetime.today().date()
    new_start_datetime = datetime.combine(today, new_class_start_time)
    new_end_datetime = datetime.combine(today, new_class_end_time)
    for existing_slot in professor_schedule[faculty][day]:
        existing_class_type = timetable[day][existing_slot]['type']
        if existing_class_type is None:
            continue
        existing_class_start_time = time_slots[existing_slot][0]
        existing_class_end_time = None
        for i in range(existing_slot, len(time_slots)):
            if i in professor_schedule[faculty][day] and timetable[day][i]['type'] is not None:
                existing_class_end_time = time_slots[i][1]
            else:
                break
        if existing_class_end_time is None:
            continue
        existing_start_datetime = datetime.combine(today, existing_class_start_time)
        existing_end_datetime = datetime.combine(today, existing_class_end_time)
        if (new_start_datetime == existing_end_datetime and 
            existing_class_type in ['LEC', 'TUT'] and duration == LAB_DURATION):
            continue
        
        if (existing_start_datetime == new_end_datetime and 
            timetable[day][start_slot]['type'] == 'LAB' and existing_class_type in ['LEC', 'TUT']):
            continue
        time_diff_hours = abs((new_start_datetime - existing_start_datetime).total_seconds() / 3600)
        if new_start_datetime == existing_end_datetime or existing_start_datetime == new_end_datetime:
            return False
        if time_diff_hours < 3:
            return False
    
    return True

def generate_all_timetables():
    
    global TIME_SLOTS
    initialize_time_slots()  
    rooms = load_rooms()
    batch_info = load_batch_data()
    faculty_preferences = load_faculty_preferences()
    reserved_slots = load_reserved_slots()
    teaching_assistants = load_teaching_assistants()
    
    wb = Workbook()
    wb.remove(wb.active)  
    overview_sheet = wb.create_sheet(title="Overview")
    overview_sheet.append(["Combined Timetable for All Departments and Semesters"])
    overview_sheet.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview_sheet.append([])
    overview_sheet.append(["Department", "Semester", "Sheet Name"])
    professor_schedule = {}
    unscheduled_components = set()
    ta_allocations = {}  # Track TA assignments {course_code: [TA names]}
    elective_schedule_sync = {}  # {(semester, course_code): (day, start_slot)}
    
    subject_colors = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1", 
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]
    basket_group_colors = {
        'B1': "FFA07A",  
        'B2': "98FB98",  
        'B3': "87CEFA",  
        'B4': "FFD700",  
        'B5': "DA70D6",  
        'B6': "20B2AA",  
        'B7': "FF6347",  
        'B8': "8A2BE2",  
        'B9': "32CD32"   
    }
    self_study_courses = []
    all_semesters = sorted(set(int(str(sem)[0]) for sem in df['Semester'].unique()))
    lunch_breaks = calculate_lunch_breaks(all_semesters)
    row_index = 5  
    
    for department in df['Department'].unique():
        course_faculty_assignments = {}
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & 
                        (df['Semester'] == semester) &
                        ((df['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                         (df['Schedule'].isna()))].copy()
            
            if courses.empty:
                continue
            elective_courses = []
            core_courses = []
            
            for _, course in courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                dept = str(course['Department'])
                if is_elective_or_minor_course(code, name, dept):
                    elective_courses.append(course)
                else:
                    core_courses.append(course)
            if elective_courses:
                elective_df = pd.DataFrame(elective_courses)
                elective_lab = elective_df[elective_df['P'] > 0].copy()
                elective_lab['priority'] = elective_lab.apply(get_course_priority, axis=1)
                elective_lab = elective_lab.sort_values('priority', ascending=False)
                elective_non_lab = elective_df[elective_df['P'] == 0].copy()
                elective_non_lab['priority'] = elective_non_lab.apply(get_course_priority, axis=1)
                elective_non_lab = elective_non_lab.sort_values('priority', ascending=False)
                sorted_electives = pd.concat([elective_lab, elective_non_lab]) if not elective_lab.empty else elective_non_lab
            else:
                sorted_electives = pd.DataFrame()
            if core_courses:
                core_df = pd.DataFrame(core_courses)
                core_lab = core_df[core_df['P'] > 0].copy()
                core_lab['priority'] = core_lab.apply(get_course_priority, axis=1)
                core_lab = core_lab.sort_values('priority', ascending=False)
                core_non_lab = core_df[core_df['P'] == 0].copy()
                core_non_lab['priority'] = core_non_lab.apply(get_course_priority, axis=1)
                core_non_lab = core_non_lab.sort_values('priority', ascending=False)
                sorted_core = pd.concat([core_lab, core_non_lab]) if not core_lab.empty else core_non_lab
            else:
                sorted_core = pd.DataFrame()
            courses = pd.concat([sorted_electives, sorted_core]).reset_index(drop=True)
            
            dept_info = batch_info.get((department, semester))
            num_sections = dept_info['num_sections'] if dept_info else 1
            for _, course in courses.iterrows():
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                p = int(course['P']) if pd.notna(course['P']) else 0
                s = int(course['S']) if pd.notna(course['S']) else 0
                
                if s > 0 and l == 0 and t == 0 and p == 0:
                    self_study_courses.append({
                        'code': str(course['Course Code']),
                        'name': str(course['Course Name']),
                        'faculty': str(course['Faculty']),
                        'department': department,
                        'semester': semester
                    })
            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                overview_sheet.cell(row=row_index, column=1, value=department)
                overview_sheet.cell(row=row_index, column=2, value=str(semester))
                overview_sheet.cell(row=row_index, column=3, value=section_title)
                row_index += 1
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
                subject_color_map = {}
                course_faculty_map = {}  
                color_idx = 0
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code and code != 'nan':
                        if is_basket_course(code):
                            basket_group = get_basket_group(code)
                            subject_color_map[code] = basket_group_colors.get(basket_group, subject_colors[color_idx % len(subject_colors)])
                        else:
                            subject_color_map[code] = subject_colors[color_idx % len(subject_colors)]
                        course_faculty_map[code] = {
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        }
                        color_idx += 1
                courses_list = list(courses.iterrows())
                elective_basket_courses = [(idx, course) for idx, course in courses_list 
                                          if is_elective_or_minor_course(str(course['Course Code']), 
                                                                        str(course['Course Name']), 
                                                                        str(course['Department']))]
                core_courses_list = [(idx, course) for idx, course in courses_list 
                                    if not is_elective_or_minor_course(str(course['Course Code']), 
                                                                       str(course['Course Name']), 
                                                                       str(course['Department']))]
                sorted_courses = elective_basket_courses + core_courses_list
                
                for _, course in sorted_courses:
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    if code.startswith('B0'):
                        print(f"DEBUG: Processing B0 course: {code} - {name}, Faculty: {faculty}, Section: {section}")
                    
                    if not any(code.startswith(f'B{i}') for i in range(1, 10)):
                        if code in course_faculty_assignments:
                            if '/' in faculty:
                                faculty_options = [f.strip() for f in faculty.split('/')] 
                                available_faculty = [f for f in faculty_options 
                                                     if f not in course_faculty_assignments[code]]
                                if available_faculty:
                                    faculty = available_faculty[0]
                                else:
                                    faculty = select_faculty(faculty)
                        else:
                            faculty = select_faculty(faculty)
                            course_faculty_assignments[code] = [faculty]
                    else:
                        faculty = select_faculty(faculty)
                    lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions = calculate_required_slots(course)
                    
                    is_elective_course_flag = is_elective_or_minor_course(code, name, department)
                    is_basket = is_basket_course(code)
                    if code.startswith('B0'):
                        print(f"DEBUG: B0 course {code} - Lecture sessions: {lecture_sessions}, Is elective: {is_elective_course_flag}, Is basket: {is_basket}")
                    
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                    for _ in range(lecture_sessions):
                        scheduled = False
                        attempts = 0
                        candidate_slots = get_candidate_start_slots(LECTURE_DURATION, is_elective_course_flag)
                        if not candidate_slots:
                            break
                        if is_basket or is_elective_course_flag:
                            sync_key = (semester, code)  # Semester-level sync for electives
                        else:
                            sync_key = (department, semester, code)  # Department-level for core courses
                        if code.startswith('B0'):
                            print(f"DEBUG: B0 course {code} section {section} - Checking sync: is_basket={is_basket}, is_elective={is_elective_course_flag}, num_sections={num_sections}, sync_key in dict={sync_key in elective_schedule_sync}")
                        if (is_basket or is_elective_course_flag) and sync_key in elective_schedule_sync:
                            day, start_slot = elective_schedule_sync[sync_key]
                            if code.startswith('B0'):
                                print(f"DEBUG: B0 course {code} section {section} - Using synchronized slot: Day {day}, Slot {start_slot}")
                            if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                scheduled = False
                                if code.startswith('B0'):
                                    print(f"DEBUG: B0 course {code} section {section} - Failed faculty_course_gap check")
                            elif not check_faculty_daily_components(professor_schedule, faculty, day, 
                                                               department, semester, section, timetable,
                                                               code, 'LEC'):
                                scheduled = False
                                if code.startswith('B0'):
                                    print(f"DEBUG: B0 course {code} section {section} - Failed faculty_daily_components check")
                            else:
                                slots_free = True
                                for i in range(LECTURE_DURATION):
                                    current_slot = start_slot + i
                                    slot_occupied = timetable[day][current_slot]['type'] is not None
                                    if slot_occupied and unscheduled_components:
                                        existing_code = timetable[day][current_slot]['code'] if i == 0 else ''
                                        if existing_code and is_elective_or_minor_course(existing_code, '', department):
                                            slot_occupied = False
                                    faculty_conflict = current_slot in professor_schedule[faculty][day] if not is_basket else False
                                    
                                    if (faculty_conflict or 
                                        slot_occupied or
                                        is_break_time(TIME_SLOTS[current_slot], semester)):
                                        slots_free = False
                                        if code.startswith('B0'):
                                            print(f"DEBUG: B0 course {code} section {section} - Sync path slots NOT free: slot={current_slot}, faculty_conflict={faculty_conflict}, slot_occupied={slot_occupied}, is_break={is_break_time(TIME_SLOTS[current_slot], semester)}")
                                        break
                                    if current_slot > 0:
                                        if is_lecture_scheduled(timetable, day, 
                                                             max(0, current_slot - BREAK_DURATION), 
                                                             current_slot):
                                            slots_free = False
                                            break
                                    if current_slot < len(TIME_SLOTS) - 1:
                                        if is_lecture_scheduled(timetable, day,
                                                             current_slot + 1,
                                                             min(len(TIME_SLOTS), 
                                                                 current_slot + BREAK_DURATION + 1)):
                                            slots_free = False
                                            break
                                
                                if slots_free:
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                              day, start_slot, LECTURE_DURATION, 
                                                              rooms, batch_info, timetable, code)
                                    
                                    if code.startswith('B0'):
                                        print(f"DEBUG: B0 course {code} section {section} - Sync path slots_free: {slots_free}, room_id: {room_id}")
                                    
                                    if room_id:
                                        classroom = room_id
                                        for i in range(LECTURE_DURATION):
                                            if not is_basket:
                                                professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LEC'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        
                                        if code.startswith('B0'):
                                            print(f"DEBUG: B0 course {code} section {section} SCHEDULED via SYNC at Day {day}, Slot {start_slot}, Room {classroom}")
                        if not scheduled:
                            if code.startswith('B0'):
                                print(f"DEBUG: B0 course {code} section {section} - Starting normal scheduling, attempts: {attempts}")
                            
                            while not scheduled and attempts < 1000:
                                day = random.randint(0, len(DAYS)-1)
                                start_slot = random.choice(candidate_slots)
                                if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                    attempts += 1
                                    continue
                                if not check_faculty_daily_components(professor_schedule, faculty, day, 
                                                                   department, semester, section, timetable,
                                                                   code, 'LEC'):
                                    attempts += 1
                                    continue
                                slots_free = True
                                for i in range(LECTURE_DURATION):
                                    current_slot = start_slot + i
                                    slot_occupied = timetable[day][current_slot]['type'] is not None
                                    if slot_occupied and unscheduled_components:
                                        existing_code = timetable[day][current_slot]['code'] if i == 0 else ''
                                        if existing_code and is_elective_or_minor_course(existing_code, '', department):
                                            slot_occupied = False
                                    
                                    if (current_slot in professor_schedule[faculty][day] or 
                                        slot_occupied or
                                        is_break_time(TIME_SLOTS[current_slot], semester)):
                                        slots_free = False
                                        break
                                    if current_slot > 0:
                                        if is_lecture_scheduled(timetable, day, 
                                                             max(0, current_slot - BREAK_DURATION), 
                                                             current_slot):
                                            slots_free = False
                                            break
                                    if current_slot < len(TIME_SLOTS) - 1:
                                        if is_lecture_scheduled(timetable, day,
                                                             current_slot + 1,
                                                             min(len(TIME_SLOTS), 
                                                                 current_slot + BREAK_DURATION + 1)):
                                            slots_free = False
                                            break
                                
                                if slots_free:
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                              day, start_slot, LECTURE_DURATION, 
                                                              rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id
                                        for i in range(LECTURE_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LEC'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        if code.startswith('B0'):
                                            print(f"DEBUG: B0 course {code} section {section} SCHEDULED at Day {day}, Slot {start_slot}, Room {classroom}")
                                            print(f"DEBUG: B0 course {code} section {section} - Checking registration: is_basket={is_basket}, is_elective={is_elective_course_flag}, sync_key in dict={sync_key in elective_schedule_sync}")
                                        if (is_basket or is_elective_course_flag) and sync_key not in elective_schedule_sync:
                                            elective_schedule_sync[sync_key] = (day, start_slot)
                                            if code.startswith('B0'):
                                                print(f"DEBUG: B0 course {code} - Registered sync schedule for semester {semester}: Day {day}, Slot {start_slot}")
                                attempts += 1
                        if not scheduled:
                                detailed_reason = unscheduled_reason(course, department, semester, 
                                                                  professor_schedule, rooms, 'LEC', attempts)
                                
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name, 
                                                       faculty, 'LEC', 1, section, detailed_reason)
                                )
                    for _ in range(tutorial_sessions):
                        scheduled = False
                        attempts = 0
                        candidate_slots = get_candidate_start_slots(TUTORIAL_DURATION, is_elective_course_flag)
                        if not candidate_slots:
                            break
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            start_slot = random.choice(candidate_slots)
                            if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                attempts += 1
                                continue
                            if not check_faculty_daily_components(professor_schedule, faculty, day,
                                                               department, semester, section, timetable,
                                                               code, 'TUT'):
                                attempts += 1
                                continue
                                
                            slots_free = True
                            for i in range(TUTORIAL_DURATION):
                                if (start_slot+i in professor_schedule[faculty][day] or 
                                    timetable[day][start_slot+i]['type'] is not None or
                                    is_break_time(TIME_SLOTS[start_slot+i], semester)):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, TUTORIAL_DURATION, 
                                                          rooms, batch_info, timetable, code)
                                
                                if room_id:
                                    classroom = room_id
                                    for i in range(TUTORIAL_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'TUT'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                            detailed_reason = unscheduled_reason(course, department, semester, 
                                                              professor_schedule, rooms, 'TUT', attempts)
                            
                            unscheduled_components.add(
                                UnscheduledComponent(department, semester, code, name,
                                                   faculty, 'TUT', 1, section, detailed_reason)
                            )
                    if lab_sessions > 0:
                        room_type = get_required_room_type(course)
                        for _ in range(lab_sessions):
                            scheduled = False
                            attempts = 0
                            scheduling_reason = ""
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            
                            for day in days:
                                possible_slots = get_best_slots(timetable, professor_schedule, 
                                                              faculty, day, LAB_DURATION, 
                                                              semester, department, is_elective_course_flag)
                                
                                for start_slot in possible_slots:
                                    room_id = find_suitable_room(room_type, department, semester,
                                                               day, start_slot, LAB_DURATION,
                                                               rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id if ',' not in str(room_id) else f"{room_id.split(',')[0]}+{room_id.split(',')[1]}"
                                        for i in range(LAB_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LAB'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        break
                                
                                if scheduled:
                                    break
                                
                            if not scheduled:
                                detailed_reason = unscheduled_reason(course, department, semester, 
                                                                  professor_schedule, rooms, 'LAB', attempts)
                                
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name,
                                                       faculty, 'LAB', 1, section, detailed_reason)
                                )
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    _, _, _, self_study_sessions = calculate_required_slots(course)
                    
                    if self_study_sessions > 0:
                        if faculty not in professor_schedule:
                            professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                        for _ in range(self_study_sessions):
                            scheduled = False
                            attempts = 0
                            candidate_slots = get_candidate_start_slots(SELF_STUDY_DURATION, is_elective_course_flag)
                            if not candidate_slots:
                                break
                            while not scheduled and attempts < 1000:
                                day = random.randint(0, len(DAYS)-1)
                                start_slot = random.choice(candidate_slots)
                                slots_free = True
                                for i in range(SELF_STUDY_DURATION):
                                    if (start_slot+i in professor_schedule[faculty][day] or 
                                        timetable[day][start_slot+i]['type'] is not None or
                                        is_break_time(TIME_SLOTS[start_slot+i], semester)):
                                        slots_free = False
                                        break
                                
                                if slots_free:
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                              day, start_slot, SELF_STUDY_DURATION, 
                                                              rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id
                                        for i in range(SELF_STUDY_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'SS'  
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                attempts += 1
                try:
                    dept_section_unsched = [c for c in unscheduled_components 
                                             if c.department == department and 
                                             c.semester == semester and
                                             (c.section == section if num_sections > 1 else True) and
                                             is_elective_or_minor_course(c.code, c.name, department)]
                    if dept_section_unsched:
                        for comp in list(dept_section_unsched):
                            duration = {'LEC': LECTURE_DURATION, 'TUT': TUTORIAL_DURATION, 'LAB': LAB_DURATION, 'SS': SELF_STUDY_DURATION}.get(comp.component_type, LECTURE_DURATION)
                            candidate_slots = get_candidate_start_slots(duration, True)
                            if not candidate_slots:
                                continue
                            scheduled_flag = False
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            for day_idx in days:
                                for start_slot in candidate_slots:
                                    if start_slot > len(TIME_SLOTS) - duration:
                                        continue
                                    faculty = comp.faculty
                                    if faculty not in professor_schedule:
                                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                                    if not check_professor_course_gap(professor_schedule, timetable, faculty, comp.code, day_idx, start_slot):
                                        continue
                                    if not check_faculty_daily_components(professor_schedule, faculty, day_idx,
                                                                          department, semester, section, timetable,
                                                                          comp.code, comp.component_type):
                                        continue
                                    slots_free = True
                                    for i in range(duration):
                                        cs = start_slot + i
                                        slot_occupied = timetable[day_idx][cs]['type'] is not None
                                        if slot_occupied and len(dept_section_unsched) > 1:
                                            existing_code = timetable[day_idx][cs]['code'] if i == 0 else ''
                                            if existing_code and is_elective_or_minor_course(existing_code, '', department):
                                                slot_occupied = False
                                        
                                        if (cs in professor_schedule[faculty][day_idx] or slot_occupied or is_break_time(TIME_SLOTS[cs], semester)):
                                            slots_free = False
                                            break
                                    if not slots_free:
                                        continue
                                    room_id = find_suitable_room('LECTURE_ROOM' if comp.component_type in ['LEC','TUT','SS'] else get_required_room_type({'P':1}),
                                                                 department, semester, day_idx, start_slot, duration,
                                                                 rooms, batch_info, timetable, comp.code)
                                    if not room_id:
                                        continue
                                    classroom = room_id
                                    for i in range(duration):
                                        professor_schedule[faculty][day_idx].add(start_slot+i)
                                        timetable[day_idx][start_slot+i]['type'] = comp.component_type
                                        timetable[day_idx][start_slot+i]['code'] = comp.code if i == 0 else ''
                                        timetable[day_idx][start_slot+i]['name'] = comp.name if i == 0 else ''
                                        timetable[day_idx][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day_idx][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    try:
                                        unscheduled_components.remove(comp)
                                    except KeyError:
                                        pass
                                    scheduled_flag = True
                                    break
                                if scheduled_flag:
                                    break
                except Exception:
                    pass
                try:
                    dept_section_any_unsched = [c for c in unscheduled_components
                                                  if c.department == department and
                                                  c.semester == semester and
                                                  (c.section == section if num_sections > 1 else True)]
                    if dept_section_any_unsched:
                        for comp in list(dept_section_any_unsched):
                            duration = {'LEC': LECTURE_DURATION, 'TUT': TUTORIAL_DURATION, 'LAB': LAB_DURATION, 'SS': SELF_STUDY_DURATION}.get(comp.component_type, LECTURE_DURATION)
                            candidate_slots = get_candidate_start_slots(duration, True)
                            if not candidate_slots:
                                continue
                            scheduled_flag = False
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            for day_idx in days:
                                for start_slot in candidate_slots:
                                    if start_slot > len(TIME_SLOTS) - duration:
                                        continue
                                    faculty = comp.faculty
                                    if faculty not in professor_schedule:
                                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                                    if not check_professor_course_gap(professor_schedule, timetable, faculty, comp.code, day_idx, start_slot):
                                        continue
                                    if not check_faculty_daily_components(professor_schedule, faculty, day_idx,
                                                                          department, semester, section, timetable,
                                                                          comp.code, comp.component_type):
                                        continue
                                    slots_free = True
                                    for i in range(duration):
                                        cs = start_slot + i
                                        if (cs in professor_schedule[faculty][day_idx] or timetable[day_idx][cs]['type'] is not None or is_break_time(TIME_SLOTS[cs], semester)):
                                            slots_free = False
                                            break
                                    if not slots_free:
                                        continue
                                    room_id = find_suitable_room('LECTURE_ROOM' if comp.component_type in ['LEC','TUT','SS'] else get_required_room_type({'P':1}),
                                                                 department, semester, day_idx, start_slot, duration,
                                                                 rooms, batch_info, timetable, comp.code)
                                    if not room_id:
                                        continue
                                    classroom = room_id
                                    for i in range(duration):
                                        professor_schedule[faculty][day_idx].add(start_slot+i)
                                        timetable[day_idx][start_slot+i]['type'] = comp.component_type
                                        timetable[day_idx][start_slot+i]['code'] = comp.code if i == 0 else ''
                                        timetable[day_idx][start_slot+i]['name'] = comp.name if i == 0 else ''
                                        timetable[day_idx][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day_idx][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    try:
                                        unscheduled_components.remove(comp)
                                    except KeyError:
                                        pass
                                    scheduled_flag = True
                                    break
                                if scheduled_flag:
                                    break
                except Exception:
                    pass
                try:
                    arrange_unscheduled_components(unscheduled_components, professor_schedule, timetable, rooms, batch_info,
                                                   department=department, semester=semester, section=section,
                                                   elective_schedule_sync=elective_schedule_sync, num_sections=num_sections)
                except Exception:
                    pass

                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                lec_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")  
                lab_fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")  
                tut_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  
                ss_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   
                break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") 
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    
                    merge_ranges = []  
                    
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        
                        if is_break_time(TIME_SLOTS[slot_idx], semester):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            if code:
                                duration = {
                                    'LEC': LECTURE_DURATION,
                                    'LAB': LAB_DURATION,
                                    'TUT': TUTORIAL_DURATION,
                                    'SS': SELF_STUDY_DURATION
                                }.get(activity_type, 1)
                                if code in subject_color_map:
                                    cell_fill = PatternFill(start_color=subject_color_map[code],
                                                          end_color=subject_color_map[code],
                                                          fill_type="solid")
                                else:
                                    cell_fill = {
                                        'LAB': lab_fill,
                                        'TUT': tut_fill,
                                        'SS': ss_fill,
                                        'LEC': lec_fill
                                    }.get(activity_type, lec_fill)
                                
                                if code and is_basket_course(code):
                                    basket_group = get_basket_group(code)
                                    basket_codes = set()  
                                    basket_details = {}
                                    for slot_id, slot_data in timetable[day_idx].items():
                                        slot_code = slot_data.get('code', '')
                                        if (slot_data.get('type') == activity_type and 
                                            get_basket_group(slot_code) == basket_group):
                                            basket_codes.add(slot_code)  
                                            if slot_code not in basket_details:
                                                basket_details[slot_code] = {
                                                    'faculty': slot_data['faculty'],
                                                    'room': slot_data['classroom']
                                                }
                                    
                                    if basket_codes:
                                        basket_header = f"{basket_group} Courses\n"
                                        codes_str = ', '.join(sorted(basket_codes))
                                        course_details = [
                                            f"{code}: {details['faculty']} ({details['room']})"
                                            for code, details in sorted(basket_details.items())
                                            if code and details['faculty'] and details['room']
                                        ]
                                        
                                        cell_value = f"{basket_header}{codes_str}\n" + "\n".join(course_details)
                                else:
                                    cell_value = f"{code} {activity_type}\nroom no. :{classroom}\n{faculty}"
                                if duration > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + duration + 1)
                                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                                    merge_ranges.append((merge_range, cell_fill))
                        
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', indent=1)
                    for merge_range, fill in merge_ranges:
                        ws.merge_cells(merge_range)
                        merged_cell = ws[merge_range.split(':')[0]]
                        merged_cell.fill = fill
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', indent=2)

                for col_idx in range(1, len(TIME_SLOTS)+2):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40
                current_row = len(DAYS) + 4  

                if self_study_courses:
                    ss_courses_for_this_section = [c for c in self_study_courses 
                                               if c['department'] == department and 
                                               c['semester'] == semester]
                    
                    if ss_courses_for_this_section:
                        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 1
                        
                        headers = ['Course Code', 'Course Name', 'Faculty']
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=current_row, column=col, value=header)
                            ws.cell(row=current_row, column=col).font = Font(bold=True)
                        current_row += 1
                        
                        for course in ss_courses_for_this_section:
                            ws.cell(row=current_row, column=1, value=course['code'])
                            ws.cell(row=current_row, column=2, value=course['name'])
                            ws.cell(row=current_row, column=3, value=course['faculty'])
                            current_row += 1
                        
                        current_row += 2  
                dept_unscheduled = [c for c in unscheduled_components 
                                    if c.department == department and 
                                    c.semester == semester and
                                    (c.section == section if num_sections > 1 else True)]

                if dept_unscheduled:
                    current_row += 2  
                    unsch_title = ws.cell(row=current_row, column=1, value="Unscheduled Components")
                    unsch_title.font = Font(bold=True, size=12, color="FF0000")
                    current_row += 2

                    headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Reason']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    current_row += 1

                    for comp in dept_unscheduled:
                        cells = [
                            (comp.code, None),
                            (comp.name, None),
                            (comp.faculty, None),
                            (comp.component_type, None),
                            (comp.sessions, None),
                            (comp.reason or "Could not find suitable slot", None)
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
                    
                    current_row += 2  
                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2
                ws.column_dimensions['A'].width = 20  
                ws.column_dimensions['B'].width = 10  
                ws.column_dimensions['C'].width = 40  
                ws.column_dimensions['D'].width = 30  
                ws.column_dimensions['E'].width = 15  
                legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                current_row += 1
                for code, color in subject_color_map.items():
                    if code in course_faculty_map:
                        ws.row_dimensions[current_row].height = 30
                        ltps_value = ""
                        for _, course_row in courses.iterrows():
                            if str(course_row['Course Code']) == code:
                                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                                ltps_value = f"{l}-{t}-{p}-{s}"
                                break
                        cells = [
                            (code, None),
                            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
                            (course_faculty_map[code]['name'], None),
                            (course_faculty_map[code]['faculty'], None),
                            (ltps_value, None)
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
                        
                        current_row += 1
    for col in range(1, 4):
        overview_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    for row in overview_sheet.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.font = Font(bold=True)
    for cell in overview_sheet[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    for row in overview_sheet.iter_rows(min_row=5, max_row=row_index-1):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
    for sheet in wb.worksheets:
        sheet.protection = SheetProtection(sheet=False, password=None)
    
    wb.save("timetable_all_departments.xlsx")
    print("Combined timetable for all departments and semesters saved as timetable_all_departments.xlsx")
    
    return ["timetable_all_departments.xlsx"]

def check_unscheduled_courses():
    
    try:
        df = pd.read_csv('combined.csv')
        if not os.path.exists('timetable_all_departments.xlsx'):
            print("Warning: timetable_all_departments.xlsx not found. Run generate_all_timetables() first.")
            return
        wb = pd.ExcelFile('timetable_all_departments.xlsx')
        scheduled_hours = defaultdict(lambda: {'L': 0, 'T': 0, 'P': 0, 'S': 0})
        found_courses = []
        
        print("\nExamining timetable for scheduled courses...")
        
        for sheet_name in wb.sheet_names:
            timetable_df = pd.read_excel(wb, sheet_name=sheet_name)
            if 'Day' not in timetable_df.columns:
                continue
            
            print(f"Processing sheet: {sheet_name}")
            for _, row in timetable_df.iterrows():
                day = row['Day']
                if day not in DAYS:  
                    continue
                for col in timetable_df.columns[1:]:  
                    cell_value = row[col]
                    if pd.isna(cell_value) or cell_value == '' or cell_value == 'BREAK':
                        continue
                    if isinstance(cell_value, str) and 'room no.' in cell_value:
                        parts = cell_value.split('room no.')
                        course_info = parts[0].strip()
                        if ' ' in course_info:
                            course_parts = course_info.split()
                            if len(course_parts) >= 2:
                                course_code = course_parts[0].strip()
                                class_type = course_parts[1].strip()
                                if 'HS204' in course_code or 'HS153' in course_code:
                                    print(f"Found in timetable: {course_code}, Type: {class_type}, Day: {day}, Slot: {col}")
                                
                                found_courses.append(course_code)  
                                if class_type == 'LEC':
                                    scheduled_hours[course_code]['L'] += 1.5  
                                    print(f"Added 1.5 lecture hours for {course_code}")
                                elif class_type == 'TUT':
                                    scheduled_hours[course_code]['T'] += 1    
                                    print(f"Added 1 tutorial hour for {course_code}")
                                elif class_type == 'LAB':
                                    scheduled_hours[course_code]['P'] += 2    
                                    print(f"Added 2 practical hours for {course_code}")
        all_found_courses = set(found_courses)
        course_primary_codes = {}  
        for _, course in df.iterrows():
            original_code = str(course['Course Code']).strip()
            if '/' in original_code:
                variants = [c.strip() for c in original_code.split('/')]
                primary = variants[0]
                for variant in variants:
                    course_primary_codes[variant] = primary
                course_primary_codes[original_code] = primary
                if 'HS204' in original_code or 'HS153' in original_code:
                    print(f"Mapping variants for {original_code}: {variants}")
            elif '(' in original_code and ')' in original_code:
                base = original_code.split('(')[0].strip()
                inner_part = original_code.split('(')[1].split(')')[0]
                
                if '/' in inner_part:
                    inner_variants = [c.strip() for c in inner_part.split('/')]
                    for variant in inner_variants:
                        if variant.lower() != 'new':
                            combined = f"{base}_{variant}"
                            course_primary_codes[combined] = original_code
                            course_primary_codes[variant] = original_code
                course_primary_codes[original_code] = original_code
            else:
                course_primary_codes[original_code] = original_code
        print("\nUnique courses found in timetable:", len(all_found_courses))
        print("First 10 courses found:", list(all_found_courses)[:10])
        merged_hours = defaultdict(lambda: {'L': 0, 'T': 0, 'P': 0, 'S': 0})
        
        for code, hours in scheduled_hours.items():
            primary_code = course_primary_codes.get(code, code)
            if 'HS204' in code or 'HS153' in code or 'HS204' in primary_code or 'HS153' in primary_code:
                print(f"Merging hours for {code} -> {primary_code}: {hours}")
            merged_hours[primary_code]['L'] += hours['L']
            merged_hours[primary_code]['T'] += hours['T']
            merged_hours[primary_code]['P'] += hours['P']
            merged_hours[primary_code]['S'] += hours['S']
        unscheduled_courses = []
        
        for _, course in df.iterrows():
            original_code = str(course['Course Code']).strip()
            name = str(course['Course Name'])
            faculty = str(course['Faculty'])
            department = str(course['Department'])
            semester = str(course['Semester'])
            primary_code = course_primary_codes.get(original_code, original_code)
            required_l = int(course['L']) if pd.notna(course['L']) else 0
            required_t = int(course['T']) if pd.notna(course['T']) else 0
            required_p = int(course['P']) if pd.notna(course['P']) else 0
            required_s = int(course['S']) if pd.notna(course['S']) and 'S' in course else 0
            scheduled_l = merged_hours[primary_code]['L']
            scheduled_t = merged_hours[primary_code]['T']
            scheduled_p = merged_hours[primary_code]['P']
            scheduled_s = merged_hours[primary_code]['S']
            if 'HS204' in original_code or 'HS153' in original_code:
                print(f"\nCourse: {original_code} (Primary: {primary_code})")
                print(f"  Required L-T-P-S: {required_l}-{required_t}-{required_p}-{required_s}")
                print(f"  Merged Scheduled L-T-P-S: {scheduled_l}-{scheduled_t}-{scheduled_p}-{scheduled_s}")
            tolerance = 0.01
            missing_l = max(0, required_l - scheduled_l)
            missing_t = max(0, required_t - scheduled_t)
            missing_p = max(0, required_p - scheduled_p)
            missing_s = max(0, required_s - scheduled_s)
            
            if (missing_l > tolerance or missing_t > tolerance or 
                missing_p > tolerance or missing_s > tolerance):
                reasons = []
                variants_found = False
                found_variants = []
                if primary_code in course_primary_codes.values():
                    for code, primary in course_primary_codes.items():
                        if primary == primary_code and code in all_found_courses:
                            variants_found = True
                            found_variants.append(code)
                else:
                    if primary_code in all_found_courses:
                        variants_found = True
                        found_variants.append(primary_code)
                
                if variants_found:
                    reasons.append(f"Course found in timetable as {', '.join(found_variants)} but not all required hours are scheduled")
                else:
                    reasons.append("Course not found in any timetable")
                faculty_courses = df[df['Faculty'] == faculty]['Course Code'].tolist()
                if len(faculty_courses) > 1:
                    reasons.append("Faculty teaching multiple courses may have scheduling constraints")
                section_key = (department, semester)
                if len(lecture_rooms) < 1 and required_l > 0:
                    reasons.append("Insufficient lecture rooms available")
                if len(computer_lab_rooms) < 1 and required_p > 0:
                    reasons.append("Insufficient lab rooms available")
                semester_courses = df[(df['Department'] == department) & (df['Semester'] == semester)].shape[0]
                if semester_courses > 6:
                    reasons.append("High number of courses in same semester may cause conflicts")
                unscheduled_courses.append({
                    'Code': original_code,
                    'Name': name,
                    'Faculty': faculty,
                    'Department': department,
                    'Semester': semester,
                    'Required L-T-P-S': f"{required_l}-{required_t}-{required_p}-{required_s}",
                    'Scheduled L-T-P-S': f"{scheduled_l}-{scheduled_t}-{scheduled_p}-{scheduled_s}",
                    'Missing L': round(missing_l, 2),
                    'Missing T': round(missing_t, 2),
                    'Missing P': round(missing_p, 2),
                    'Missing S': round(missing_s, 2),
                    'Variant Found': variants_found,
                    'Found As': ', '.join(found_variants) if found_variants else "Not found",
                    'Reasons': "; ".join(reasons)
                })
        if unscheduled_courses:
            print("\n=== COURSES WITH UNSCHEDULED HOURS ===")
            print(f"Found {len(unscheduled_courses)} courses with scheduling issues:\n")
            
            for course in unscheduled_courses:
                print(f"Course: {course['Code']} - {course['Name']}")
                print(f"  Department: {course['Department']}, Semester: {course['Semester']}")
                print(f"  Faculty: {course['Faculty']}")
                print(f"  Required L-T-P-S: {course['Required L-T-P-S']}")
                print(f"  Scheduled L-T-P-S: {course['Scheduled L-T-P-S']}")
                
                missing = []
                if course['Missing L'] > 0:
                    missing.append(f"{course['Missing L']} lecture hours")
                if course['Missing T'] > 0:
                    missing.append(f"{course['Missing T']} tutorial hours")
                if course['Missing P'] > 0:
                    missing.append(f"{course['Missing P']} practical hours")
                if course['Missing S'] > 0:
                    missing.append(f"{course['Missing S']} self-study hours")
                    
                print(f"  Missing: {', '.join(missing)}")
                print(f"  Found in Timetable: {'Yes' if course['Variant Found'] else 'No'}")
                if course['Variant Found']:
                    print(f"  Found as: {course['Found As']}")
                print(f"  Possible Reasons: {course['Reasons']}\n")
            unscheduled_df = pd.DataFrame(unscheduled_courses)
            unscheduled_df.to_excel('unscheduled_courses.xlsx', index=False)
            print("Details saved to 'unscheduled_courses.xlsx'")
        else:
            print("\n=== ALL COURSES FULLY SCHEDULED ===")
            print("All courses have been scheduled according to their L-T-P-S requirements.")
            
    except Exception as e:
        print(f"Error checking unscheduled courses: {e}")
        import traceback
        traceback.print_exc()

def generate_faculty_timetables():
    
    try:
        wb = pd.ExcelFile('timetable_all_departments.xlsx')
        faculty_schedules = {}
        
        print("Processing timetable sheets to extract faculty schedules...")
        try:
            courses_data = pd.read_csv('combined.csv')
        except Exception as e:
            print(f"Warning: Could not load combined.csv for reference: {e}")
            courses_data = None
        for sheet_name in wb.sheet_names:
            timetable_df = pd.read_excel(wb, sheet_name=sheet_name)
            if 'Day' not in timetable_df.columns:
                continue
                
            print(f"Processing sheet: {sheet_name}")
            dept_sem = sheet_name
            for _, row in timetable_df.iterrows():
                day = row['Day']
                if day not in DAYS:  
                    continue
                for col in timetable_df.columns[1:]:  
                    cell_value = row[col]
                    if pd.isna(cell_value) or cell_value == '' or cell_value == 'BREAK':
                        continue
                    if isinstance(cell_value, str) and "room no." in cell_value:
                        try:
                            lines = cell_value.strip().split('\n')
                            first_line_parts = lines[0].split()
                            if len(first_line_parts) >= 2:
                                course_code = first_line_parts[0]
                                class_type = first_line_parts[1]
                                room_info = lines[1].replace('room no. :', '').strip() if len(lines) > 1 else "Unknown"
                                faculty_name = lines[2].strip() if len(lines) > 2 else ""
                                if not faculty_name or len(faculty_name) < 2:
                                    if courses_data is not None:
                                        course_row = courses_data[courses_data['Course Code'] == course_code]
                                        if not course_row.empty:
                                            faculty_name = str(course_row['Faculty'].iloc[0])
                                course_name = ""
                                if courses_data is not None:
                                    course_row = courses_data[courses_data['Course Code'] == course_code]
                                    if not course_row.empty:
                                        course_name = str(course_row['Course Name'].iloc[0])
                                if faculty_name:
                                    faculty_list = extract_faculty_names(faculty_name)
                                    
                                    for faculty in faculty_list:
                                        if faculty.strip():
                                            canonical_name = get_canonical_faculty_name(faculty_schedules, faculty)
                                            
                                            if canonical_name not in faculty_schedules:
                                                faculty_schedules[canonical_name] = {d: {} for d in DAYS}
                                            time_slot_str = col
                                            faculty_schedules[canonical_name][day][time_slot_str] = {
                                                'Course Code': course_code,
                                                'Course Name': course_name,
                                                'Class Type': class_type,
                                                'Room': room_info,
                                                'Department-Semester': dept_sem
                                            }
                        except Exception as e:
                            print(f"Error processing cell: {cell_value}")
                            print(f"Error details: {e}")
                            traceback.print_exc()
                    elif isinstance(cell_value, str) and "Courses" in cell_value:
                        try:
                            lines = cell_value.strip().split('\n')
                            basket_courses = []
                            basket_details = []
                            for line in lines[1:]:
                                if ':' in line:  
                                    parts = line.split(':')
                                    if len(parts) >= 2:
                                        code = parts[0].strip()
                                        details = parts[1].strip()
                                        if '(' in details and ')' in details:
                                            faculty_part = details.split('(')[0].strip()
                                            room_part = details.split('(')[1].split(')')[0].strip()
                                            
                                            basket_details.append({
                                                'code': code,
                                                'faculty': faculty_part,
                                                'room': room_part
                                            })
                                elif ',' in line and not any(x in line for x in ['Courses', 'room no.']):
                                    basket_courses = [code.strip() for code in line.split(',')]
                            for detail in basket_details:
                                code = detail['code']
                                faculty_name = detail['faculty']
                                room_info = detail['room']
                                course_name = ""
                                if courses_data is not None:
                                    course_row = courses_data[courses_data['Course Code'] == code]
                                    if not course_row.empty:
                                        course_name = str(course_row['Course Name'].iloc[0])
                                class_type = "LEC"  
                                if "LAB" in cell_value:
                                    class_type = "LAB"
                                elif "TUT" in cell_value:
                                    class_type = "TUT"
                                faculty_list = extract_faculty_names(faculty_name)
                                
                                for faculty in faculty_list:
                                    if faculty.strip():
                                        canonical_name = get_canonical_faculty_name(faculty_schedules, faculty)
                                        
                                        if canonical_name not in faculty_schedules:
                                            faculty_schedules[canonical_name] = {d: {} for d in DAYS}
                                        time_slot_str = col
                                        faculty_schedules[canonical_name][day][time_slot_str] = {
                                            'Course Code': code,
                                            'Course Name': course_name,
                                            'Class Type': class_type,
                                            'Room': room_info,
                                            'Department-Semester': dept_sem
                                        }
                        except Exception as e:
                            print(f"Error processing basket cell: {cell_value}")
                            print(f"Error details: {e}")
        print(f"Creating a consolidated Excel file with {len(faculty_schedules)} faculty timetables...")
        faculty_wb = Workbook()
        if "Sheet" in faculty_wb.sheetnames:
            faculty_wb.remove(faculty_wb["Sheet"])
        overview = faculty_wb.create_sheet("Overview", 0)
        overview.column_dimensions['A'].width = 40
        overview.column_dimensions['B'].width = 15
        overview.append(["Faculty Timetable - All Faculty"])
        overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        overview.append([])
        overview.append(["Faculty Name", "Total Classes"])
        for row in range(1, 5):
            for cell in overview[row]:
                cell.font = Font(bold=True)
        for cell in overview[4]:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_idx = 5
        for faculty in sorted(faculty_schedules.keys()):
            total_classes = len(sum([list(slots.keys()) for slots in faculty_schedules[faculty].values()], []))
            overview.cell(row=row_idx, column=1, value=faculty)
            overview.cell(row=row_idx, column=2, value=total_classes)
            safe_name = sanitize_sheet_name(faculty)
            overview.cell(row=row_idx, column=1).hyperlink = f"#{safe_name}!A1"
            overview.cell(row=row_idx, column=1).style = "Hyperlink"
            for col in range(1, 3):
                overview.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            row_idx += 1
        for i, faculty in enumerate(sorted(faculty_schedules.keys())):
            sheet_name = sanitize_sheet_name(faculty)
            ws = faculty_wb.create_sheet(title=sheet_name)
            create_faculty_worksheet(ws, faculty, faculty_schedules[faculty])
            
            if i % 10 == 0:  
                print(f"Generated {i+1}/{len(faculty_schedules)} faculty worksheets")
        for sheet in faculty_wb.worksheets:
            sheet.protection = SheetProtection(sheet=False, password=None)
        
        faculty_wb.save("all_faculty_timetables.xlsx")
        print(f"All {len(faculty_schedules)} faculty timetables saved in 'all_faculty_timetables.xlsx'")
        
    except Exception as e:
        print(f"Error generating faculty timetables: {e}")
        traceback.print_exc()

def sanitize_sheet_name(name):
    
    invalid_chars = ['/', '\\', '?', '*', ':', '[', ']', "'", '"']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    if len(sanitized) > 31:
        sanitized = sanitized[:28] + "..."
    
    return sanitized

def create_faculty_worksheet(ws, faculty, schedule):
    """Create faculty timetable worksheet in grid format matching department timetable style"""
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Schedule for: {faculty}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    lec_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")  # Salmon
    lab_fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")  # Green
    tut_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Sky Blue
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") # Gray
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    for day_idx, day in enumerate(DAYS):
        row_num = day_idx + 3
        ws.append([day])
        day_schedule = schedule[day]
        slot_activities = {}  # slot_idx -> class_info
        for time_slot_str, class_info in day_schedule.items():
            for slot_idx, (start, end) in enumerate(TIME_SLOTS):
                slot_str = f"{start.strftime('%H:%M')}-{end.strftime('%H:%M')}"
                if time_slot_str == slot_str or time_slot_str.startswith(start.strftime('%H:%M')):
                    slot_activities[slot_idx] = class_info
                    break
        
        merge_ranges = []
        processed_slots = set()
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_value = ''
            cell_fill = None
            if slot_idx in processed_slots:
                continue
            if slot_idx in slot_activities:
                class_info = slot_activities[slot_idx]
                code = class_info['Course Code']
                course_name = class_info['Course Name']
                class_type = class_info['Class Type']
                room = class_info['Room']
                dept_sem = class_info['Department-Semester']
                duration = {
                    'LEC': LECTURE_DURATION,
                    'LAB': LAB_DURATION,
                    'TUT': TUTORIAL_DURATION,
                }.get(class_type, 1)
                cell_fill = {
                    'LAB': lab_fill,
                    'TUT': tut_fill,
                    'LEC': lec_fill
                }.get(class_type, lec_fill)
                cell_value = f"{code} {class_type}\n{course_name}\nRoom: {room}\n{dept_sem}"
                for i in range(slot_idx, min(slot_idx + duration, len(TIME_SLOTS))):
                    processed_slots.add(i)
                if duration > 1:
                    start_col = get_column_letter(slot_idx + 2)
                    end_col = get_column_letter(min(slot_idx + duration + 1, len(TIME_SLOTS) + 1))
                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                    merge_ranges.append((merge_range, cell_fill))
            cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
            if cell_fill:
                cell.fill = cell_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        for merge_range, fill in merge_ranges:
            ws.merge_cells(merge_range)
            merged_cell = ws[merge_range.split(':')[0]]
            merged_cell.fill = fill
            merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    for col_idx in range(1, len(TIME_SLOTS)+2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18
    for row in ws.iter_rows(min_row=3, max_row=len(DAYS)+2):
        ws.row_dimensions[row[0].row].height = 60
def extract_faculty_names(faculty_string):
    
    if not faculty_string or pd.isna(faculty_string):
        return []
        
    faculty_string = str(faculty_string).strip()
    if faculty_string.lower() in ['nan', 'none', '']:
        return []
        
    faculty_names = []
    if '&' in faculty_string:
        parts = faculty_string.split('&')
        for part in parts:
            faculty_names.append(part.strip())
    elif ' and ' in faculty_string.lower():
        parts = faculty_string.lower().split(' and ')
        for i, part in enumerate(parts):
            start_idx = faculty_string.lower().find(part)
            if start_idx >= 0:
                end_idx = start_idx + len(part)
                faculty_names.append(faculty_string[start_idx:end_idx].strip())
    elif ',' in faculty_string and faculty_string.count(',') > 1:
        parts = faculty_string.split(',')
        for part in parts:
            if part.strip():  
                faculty_names.append(part.strip())
    elif '/' in faculty_string:
        parts = faculty_string.split('/')
        for part in parts:
            faculty_names.append(part.strip())
    elif ';' in faculty_string:
        parts = faculty_string.split(';')
        for part in parts:
            faculty_names.append(part.strip())
    else:
        faculty_names.append(faculty_string)  
    return [name.strip() for name in faculty_names if name.strip()]

def normalize_faculty_name(name):
    """Normalize faculty name to handle spelling variations
    This helps match names like 'Anusree' with 'Anushree', etc.
    """
    if not name:
        return ""
    normalized = name.lower().strip()
    normalized = ' '.join(normalized.split())
    variations = {
        'anusree': 'anushree',
        'anushri': 'anushree',
        'jagadeesha': 'jagadeesha',
        'jagadeesh': 'jagadeesha',
    }
    for variant, standard in variations.items():
        if variant in normalized:
            normalized = normalized.replace(variant, standard)
    
    return normalized

def get_canonical_faculty_name(faculty_names_dict, new_name):
    """Find the canonical (existing) name for a faculty member
    Returns existing name if found (case-insensitive match), otherwise returns new_name
    """
    new_normalized = normalize_faculty_name(new_name)
    
    for existing_name in faculty_names_dict.keys():
        existing_normalized = normalize_faculty_name(existing_name)
        if new_normalized == existing_normalized:
            return existing_name
    
    return new_name

def generate_individual_faculty_timetable(faculty, schedule):
    """Generate individual faculty timetable file in grid format"""
    
    filename = sanitize_filename(faculty)
    file_path = os.path.join('faculty_timetables', f"timetable_{filename}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    create_faculty_worksheet(ws, faculty, schedule)
    
    try:
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Error saving individual timetable for {faculty}: {e}")
        return False

def sanitize_filename(name):
    
    invalid_chars = ['/', '\\', '?', '*', ':', '[', ']', "'", '"', '<', '>', '|', ' ']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    while '__' in sanitized:
        sanitized = sanitized.replace('__', '_')
    if len(sanitized) > 50:
        sanitized = sanitized[:50]
    sanitized = sanitized.rstrip('_')
    
    return sanitized
TIME_SLOTS = []
lunch_breaks = {}  

def get_faculty_preferred_slot(faculty, course_code, faculty_preferences):
    """Get the preferred day and time slot for a faculty-course combination"""
    key = (faculty, course_code)
    if key in faculty_preferences:
        prefs = faculty_preferences[key]
        priority_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
        sorted_prefs = sorted(prefs, key=lambda x: priority_order.get(x['priority'], 3))
        
        for pref in sorted_prefs:
            try:
                day_idx = DAYS.index(pref['day'])
                time_obj = datetime.strptime(pref['time'], '%H:%M').time()
                for slot_idx, (slot_start, slot_end) in enumerate(TIME_SLOTS):
                    if slot_start == time_obj:
                        return (day_idx, slot_idx, pref['priority'])
            except (ValueError, KeyError):
                continue
    
    return None

def allocate_teaching_assistants(course_code, course_name, faculty, total_students, department, tas, ta_allocations, threshold=100):
    """Allocate TAs to courses with more than threshold students"""
    if total_students < threshold:
        return []
    if course_code in ta_allocations:
        return ta_allocations[course_code]
    
    allocated_tas = []
    suitable_tas = []
    for ta in tas:
        if ta['department'] != department:
            continue
        if course_code in ta['course_preferences'] or 'ALL' in ta['course_preferences']:
            suitable_tas.append(ta)
    num_tas_needed = max(1, total_students // 50)
    
    for i, ta in enumerate(suitable_tas):
        if i >= num_tas_needed:
            break
        allocated_tas.append(ta['name'])
    
    if allocated_tas:
        ta_allocations[course_code] = allocated_tas
        print(f"Allocated TAs to {course_code}: {', '.join(allocated_tas)}")
    
    return allocated_tas

def load_config():
    
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            return config['duration_constants']
    except:
        return {
            'hour_slots': 2,
            'lecture_duration': 3,
            'lab_duration': 4,
            'tutorial_duration': 2,
            'self_study_duration': 2, 
            'break_duration': 1
        }

def initialize_time_slots():
    
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()
    initialize_elective_slots()

def initialize_elective_slots():
    
    global ELECTIVE_SLOT_INDICES, ELECTIVE_SLOT_SET, ELECTIVE_SLOT_START_TIMES
    global ELECTIVE_USE_LAST
    ELECTIVE_SLOT_INDICES = []
    try:
        with open('config.json', 'r') as f:
            cfg = json.load(f)
            slots = cfg.get('elective_slots')
            if slots and isinstance(slots, list):
                parsed = []
                for s in slots:
                    if not isinstance(s, str):
                        continue
                    try:
                        hh, mm = s.split(':')
                        parsed.append(time(int(hh), int(mm)))
                    except Exception:
                        try:
                            parsed.append(time.fromisoformat(s))
                        except Exception:
                            continue
                if parsed:
                    ELECTIVE_SLOT_START_TIMES = parsed
    except Exception:
        pass
    ELECTIVE_USE_LAST = False
    for idx, slot in enumerate(TIME_SLOTS):
        slot_start = slot[0]
        if slot_start in ELECTIVE_SLOT_START_TIMES:
            ELECTIVE_SLOT_INDICES.append(idx)
    try:
        with open('config.json', 'r') as f:
            cfg = json.load(f)
            slots = cfg.get('elective_slots')
            if slots and isinstance(slots, list):
                for s in slots:
                    if isinstance(s, str) and s.strip().upper() == 'LAST':
                        ELECTIVE_USE_LAST = True
                        break
    except Exception:
        pass
    ELECTIVE_SLOT_SET = set(ELECTIVE_SLOT_INDICES)

def get_candidate_start_slots(duration, is_elective):
    
    max_start = len(TIME_SLOTS) - duration
    if max_start < 0:
        return []
    if is_elective:
        candidates = []
        if ELECTIVE_SLOT_INDICES:
            candidates.extend([slot for slot in ELECTIVE_SLOT_INDICES if slot <= max_start])
        if 'ELECTIVE_USE_LAST' in globals() and ELECTIVE_USE_LAST:
            if max_start >= 0 and max_start not in candidates:
                candidates.append(max_start)
        if candidates:
            return sorted(set(candidates))
        return list(range(max_start + 1))
    non_elective_slots = []
    for slot in range(max_start + 1):
        if slot in ELECTIVE_SLOT_SET:
            continue
        try:
            slot_start_time = TIME_SLOTS[slot][0]
        except Exception:
            slot_start_time = None
        if slot_start_time is None:
            continue
        if slot_start_time >= time(11, 0):
            non_elective_slots.append(slot)
    if non_elective_slots:
        return non_elective_slots
    return [slot for slot in range(max_start + 1) if slot not in ELECTIVE_SLOT_SET]

def calculate_lunch_breaks(semesters):
    
    global lunch_breaks
    lunch_breaks = {}  
    if not LUNCH_WINDOW_START or not LUNCH_WINDOW_END:
        return lunch_breaks
    
    total_semesters = len(semesters)
    
    if total_semesters == 0:
        return lunch_breaks
    
    total_window_minutes = (
        LUNCH_WINDOW_END.hour * 60 + LUNCH_WINDOW_END.minute -
        LUNCH_WINDOW_START.hour * 60 - LUNCH_WINDOW_START.minute
    )
    stagger_interval = (total_window_minutes - LUNCH_DURATION) / (total_semesters - 1) if total_semesters > 1 else 0
    sorted_semesters = sorted(semesters)
    
    for i, semester in enumerate(sorted_semesters):
        start_minutes = (LUNCH_WINDOW_START.hour * 60 + LUNCH_WINDOW_START.minute + 
                        int(i * stagger_interval))
        start_hour = start_minutes // 60
        start_min = start_minutes % 60
        
        end_minutes = start_minutes + LUNCH_DURATION
        end_hour = end_minutes // 60
        end_min = end_minutes % 60
        
        lunch_breaks[semester] = (
            time(start_hour, start_min),
            time(end_hour, end_min)
        )
    
    return lunch_breaks

def is_break_time(slot, semester=None):
    
    global lunch_breaks
    start, end = slot
    morning_break = False
    if CONFIG.get('break_settings', {}).get('morning_break', {}).get('enabled', False):
        morning_start = parse_time_config(CONFIG['break_settings']['morning_break']['start_time'])
        morning_duration = CONFIG['break_settings']['morning_break']['duration_minutes']
        morning_end_dt = datetime.combine(datetime.today(), morning_start) + timedelta(minutes=morning_duration)
        morning_end = morning_end_dt.time()
        morning_break = (morning_start <= start < morning_end)
    lunch_break = False
    if semester:
        base_sem = int(str(semester)[0])  
        if base_sem in lunch_breaks:
            lunch_start, lunch_end = lunch_breaks[base_sem]
            lunch_break = (lunch_start <= start < lunch_end)
    else:
        if LUNCH_WINDOW_START and LUNCH_WINDOW_END:
            lunch_break = (LUNCH_WINDOW_START <= start < LUNCH_WINDOW_END)
        else:
            lunch_break = any(lunch_start <= start < lunch_end 
                             for lunch_start, lunch_end in lunch_breaks.values())
    
    return morning_break or lunch_break

def load_rooms():
    
    rooms = {}
    try:
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(DAYS))}
                }
    except FileNotFoundError:
        print("Warning: rooms.csv not found, using default room allocation")
        return None
    return rooms

def load_batch_data():
    
    batch_info = {}
    try:
        df = pd.read_csv('combined.csv')
        grouped = df.groupby(['Department', 'Semester'])
        
        for (dept, sem), group in grouped:
            if 'total_students' in group.columns:
                valid_students = []
                for val in group['total_students']:
                    try:
                        if pd.notna(val) and str(val).isdigit():
                            valid_students.append(int(val))
                    except (ValueError, TypeError):
                        continue
                
                if valid_students:
                    total_students = max(valid_students)
                    max_batch_size = 70
                    num_sections = (total_students + max_batch_size - 1) // max_batch_size
                    section_size = (total_students + num_sections - 1) // num_sections

                    batch_info[(dept, sem)] = {
                        'total': total_students,
                        'num_sections': num_sections,
                        'section_size': section_size
                    }
        basket_courses = df[df['Course Code'].astype(str).str.contains('^B[0-9]')]
        for _, course in basket_courses.iterrows():
            code = str(course['Course Code'])
            if 'total_students' in df.columns:
                try:
                    val = course['total_students']
                    if pd.notna(val) and str(val).isdigit():
                        total_students = int(val)
                    else:
                        total_students = 35
                except (ValueError, TypeError):
                    total_students = 35
            else:
                total_students = 35
                
            batch_info[('ELECTIVE', code)] = {
                'total': total_students,
                'num_sections': 1,  
                'section_size': total_students
            }
            
    except FileNotFoundError:
        print("Warning: combined.csv not found, using default batch sizes")
    except Exception as e:
        print(f"Warning: Error processing batch data from combined.csv: {e}")
        
    return batch_info

def is_basket_course(code):
    
    return code.startswith('B') and '-' in code

def is_minor_course(code, name=None):
    
    code_str = str(code).upper() if code is not None else ""
    name_str = str(name).upper() if name is not None else ""
    return 'MINOR' in code_str or 'MINOR' in name_str

def is_elective_or_minor_course(code, name='', department=''):
    
    code_str = str(code).upper()
    dept_str = str(department).upper()
    name_str = str(name).upper()
    if is_basket_course(code_str):
        return True
    if 'ELECTIVE' in dept_str or 'ELECTIVE' in name_str:
        return True
    if is_minor_course(code_str, name_str):
        return True
    return False

def get_basket_group(code):
    
    if is_basket_course(code):
        return code.split('-')[0]
    return None

def get_basket_group_slots(timetable, day, basket_group):
    
    basket_slots = []
    for slot_idx, slot in timetable[day].items():
        code = slot.get('code', '')
        if code and get_basket_group(code) == basket_group:
            basket_slots.append(slot_idx)
    return basket_slots

def find_adjacent_lab_room(room_id, rooms):
    
    if not room_id:
        return None
    current_num = int(''.join(filter(str.isdigit, rooms[room_id]['roomNumber'])))
    current_floor = current_num // 100
    for rid, room in rooms.items():
        if rid != room_id and room['type'] == rooms[room_id]['type']:
            room_num = int(''.join(filter(str.isdigit, room['roomNumber'])))
            if room_num // 100 == current_floor and abs(room_num - current_num) == 1:
                return rid
    return None

def try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids):
    
    for room_id, room in rooms.items():
        if room_id in used_room_ids or room['type'].upper() == 'LIBRARY':
            continue
        if course_type in ['LEC', 'TUT', 'SS']:
            if not ('LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()):
                continue
        elif course_type == 'COMPUTER_LAB' and room['type'].upper() != 'COMPUTER_LAB':
            continue
        elif course_type == 'HARDWARE_LAB' and room['type'].upper() != 'HARDWARE_LAB':
            continue
        if course_type not in ['COMPUTER_LAB', 'HARDWARE_LAB'] and room['capacity'] < required_capacity:
            continue
        slots_free = True
        for i in range(duration):
            if start_slot + i in room['schedule'][day]:
                slots_free = False
                break
                
        if slots_free:
            for i in range(duration):
                room['schedule'][day].add(start_slot + i)
            return room_id
                
    return None

def get_required_room_type(course):
    
    if pd.notna(course['P']) and course['P'] > 0:
        course_code = str(course['Course Code']).upper()
        if 'CS' in course_code or 'DS' in course_code:
            return 'COMPUTER_LAB'
        elif 'EC' in course_code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'  
    else:
        return 'LECTURE_ROOM'

def find_suitable_room(course_type, department, semester, day, start_slot, duration, rooms, batch_info, timetable, course_code="", used_rooms=None):
    
    if not rooms:
        return "DEFAULT_ROOM"
    
    required_capacity = 60  
    is_basket = is_basket_course(course_code)
    total_students = None
    
    try:
        df = pd.read_csv('combined.csv')
        
        if course_code and not is_basket:
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns:
                total_students = int(course_row['total_students'].iloc[0])
        elif is_basket:
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns:
                total_students = int(course_row['total_students'].iloc[0])
            else:
                elective_info = batch_info.get(('ELECTIVE', course_code))
                if elective_info:
                    total_students = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                total_students = dept_info['section_size']
    except Exception as e:
        print(f"Warning: Error getting total_students from combined.csv: {e}")
    if total_students:
        required_capacity = total_students
    elif batch_info:
        if is_basket:
            elective_info = batch_info.get(('ELECTIVE', course_code))
            if elective_info:
                required_capacity = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                required_capacity = dept_info['section_size']

    used_room_ids = set() if used_rooms is None else used_rooms
    if course_type in ['LEC', 'TUT', 'SS'] and required_capacity > 70:
        seater_120_rooms = {rid: room for rid, room in rooms.items() 
                           if 'SEATER_120' in room['type'].upper()}
        if required_capacity > 120:
            seater_240_rooms = {rid: room for rid, room in rooms.items() 
                              if 'SEATER_240' in room['type'].upper()}
            room_id = try_room_allocation(seater_240_rooms, 'LEC', required_capacity,
                                        day, start_slot, duration, used_room_ids)
            if room_id:
                return room_id
        room_id = try_room_allocation(seater_120_rooms, 'LEC', required_capacity,
                                    day, start_slot, duration, used_room_ids)
        if room_id:
            return room_id
    if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
        if required_capacity > 35:  
            for room_id, room in rooms.items():
                if room_id in used_room_ids or room['type'].upper() != course_type:
                    continue
                slots_free = True
                for i in range(duration):
                    if start_slot + i in room['schedule'][day]:
                        slots_free = False
                        break
                
                if slots_free:
                    adjacent_room = find_adjacent_lab_room(room_id, rooms)
                    if adjacent_room and adjacent_room not in used_room_ids:
                        adjacent_free = True
                        for i in range(duration):
                            if start_slot + i in rooms[adjacent_room]['schedule'][day]:
                                adjacent_free = False
                                break
                        
                        if adjacent_free:
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                                rooms[adjacent_room]['schedule'][day].add(start_slot + i)
                            return f"{room_id},{adjacent_room}"  
        return try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)
    if course_type in ['LEC', 'TUT', 'SS'] or is_basket:
        lecture_rooms = {rid: room for rid, room in rooms.items() 
                        if 'LECTURE_ROOM' in room['type'].upper()}
        if is_basket:
            basket_group = get_basket_group(course_code)
            basket_used_rooms = set()
            basket_group_rooms = {}  
            room_usage = {rid: sum(len(room['schedule'][d]) for d in range(len(DAYS))) 
                         for rid, room in rooms.items()}
            sorted_lecture_rooms = dict(sorted(lecture_rooms.items(), 
                                             key=lambda x: room_usage[x[0]]))
            for room_id, room in sorted_lecture_rooms.items():
                is_used = False
                for slot in range(start_slot, start_slot + duration):
                    if slot in rooms[room_id]['schedule'][day]:
                        if slot in timetable[day]:
                            slot_data = timetable[day][slot]
                            if (slot_data['classroom'] == room_id and 
                                slot_data['type'] is not None):
                                slot_code = slot_data.get('code', '')
                                if get_basket_group(slot_code) == basket_group:
                                    basket_group_rooms[slot_code] = room_id
                                else:
                                    basket_used_rooms.add(room_id)
                        is_used = True
                        break
                if not is_used and room_id not in basket_used_rooms:
                    if 'capacity' in room and room['capacity'] >= required_capacity:
                        for i in range(duration):
                            room['schedule'][day].add(start_slot + i)
                        return room_id
            if course_code in basket_group_rooms:
                return basket_group_rooms[course_code]
            room_id = try_room_allocation(lecture_rooms, 'LEC', required_capacity,
                                        day, start_slot, duration, basket_used_rooms)
            
            if room_id:
                basket_group_rooms[course_code] = room_id
            
            return room_id
        return try_room_allocation(lecture_rooms, 'LEC', required_capacity,
                                 day, start_slot, duration, used_room_ids)
    return try_room_allocation(rooms, course_type, required_capacity,
                             day, start_slot, duration, used_room_ids)

def check_faculty_daily_components(professor_schedule, faculty, day, department, semester, section, timetable, course_code=None, activity_type=None):
    
    component_count = 0
    faculty_courses = set()  
    for slot in timetable[day].values():
        if slot['faculty'] == faculty and slot['type'] in ['LEC', 'LAB', 'TUT']:
            slot_code = slot.get('code', '')
            if slot_code:
                if not is_basket_course(slot_code):
                    component_count += 1
                elif slot_code not in faculty_courses:
                    component_count += 1
                    faculty_courses.add(slot_code)
    if course_code and is_basket_course(course_code):
        basket_group = get_basket_group(course_code)
        existing_slots = get_basket_group_slots(timetable, day, basket_group)
        if existing_slots:
            return component_count < 3  
    
    return component_count < 2  

def check_faculty_course_gap(professor_schedule, timetable, faculty, course_code, day, start_slot):
    
    min_gap_hours = 3
    slots_per_hour = 2  
    required_gap = min_gap_hours * slots_per_hour
    for i in range(max(0, start_slot - required_gap), start_slot):
        if i in professor_schedule[faculty][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
    for i in range(start_slot + 1, min(len(TIME_SLOTS), start_slot + required_gap)):
        if i in professor_schedule[faculty][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
    
    return True

def is_preferred_slot(faculty, day, time_slot, faculty_preferences):
    
    if faculty not in faculty_preferences:
        return True  
        
    prefs = faculty_preferences[faculty]
    if prefs['preferred_days'] and DAYS[day] not in prefs['preferred_days']:
        return False
    if prefs['preferred_times']:
        slot_start, slot_end = time_slot
        for pref_start, pref_end in prefs['preferred_times']:
            if (slot_start >= pref_start and slot_end <= pref_end):
                return True
        return False
        
    return True  

def is_lecture_scheduled(timetable, day, start_slot, end_slot):
    
    for slot in range(start_slot, end_slot):
        if (slot < len(timetable[day]) and 
            timetable[day][slot]['type'] and 
            timetable[day][slot]['type'] in ['LEC', 'LAB', 'TUT']):
            return True
    return False

def get_best_slots(timetable, professor_schedule, faculty, day, duration, semester, department, is_elective=False):
    
    best_slots = []
    
    for start_slot in range(len(TIME_SLOTS) - duration + 1):
        if is_elective and ELECTIVE_SLOT_SET and start_slot not in ELECTIVE_SLOT_SET:
            continue
        if not is_elective and ELECTIVE_SLOT_SET and start_slot in ELECTIVE_SLOT_SET:
            continue
        slots_free = True
        for i in range(duration):
            current_slot = start_slot + i
            if duration == LAB_DURATION:
                if (current_slot in professor_schedule[faculty][day] or
                    timetable[day][current_slot]['type'] is not None or  
                    is_break_time(TIME_SLOTS[current_slot], semester)):
                    slots_free = False
                    break
            else:
                if (current_slot in professor_schedule[faculty][day] or
                    (timetable[day][current_slot]['type'] is not None and
                     not is_basket_course(timetable[day][current_slot].get('code', ''))) or
                    is_break_time(TIME_SLOTS[current_slot], semester)):
                    slots_free = False
                    break

        if slots_free:
            if duration == LAB_DURATION:
                slot_time = TIME_SLOTS[start_slot][0]
                if slot_time < time(12, 30):  
                    best_slots.insert(0, start_slot)  
                else:
                    best_slots.append(start_slot)
            else:
                best_slots.append(start_slot)
    
    return best_slots

def is_slot_reserved(slot, day, semester, department, reserved_slots):
    """Check if a time slot is reserved by coordinator"""
    if not reserved_slots or day not in reserved_slots:
        return False
        
    slot_start_time, slot_end_time = TIME_SLOTS[slot]
    
    for (dept, semesters), time_ranges in reserved_slots[day].items():
        if dept != 'ALL' and dept != department:
            continue
        if 'ALL' not in semesters:
            if str(semester) not in semesters and not any(str(semester).startswith(s) for s in semesters):
                continue
        for time_range in time_ranges:
            reserved_start = datetime.strptime(time_range[0], '%H:%M').time()
            reserved_end = datetime.strptime(time_range[1], '%H:%M').time()
            
            if (slot_start_time >= reserved_start and slot_start_time < reserved_end) or \
               (slot_end_time > reserved_start and slot_end_time <= reserved_end):
                return True
    
    return False

def get_course_priority(course):
    
    priority = 0
    code = str(course['Course Code'])
    if pd.notna(course['P']) and course['P'] > 0 and not is_basket_course(code):
        priority += 10  
        if 'CS' in code or 'EC' in code:  
            priority += 2
    elif is_basket_course(code):
        priority += 1  
    elif pd.notna(course['L']) and course['L'] > 2:
        priority += 3  
    elif pd.notna(course['T']) and course['T'] > 0:
        priority += 2  
    return priority

def calculate_required_slots(course):
    
    l = float(course['L']) if pd.notna(course['L']) else 0  
    t = int(course['T']) if pd.notna(course['T']) else 0    
    p = int(course['P']) if pd.notna(course['P']) else 0    
    s = int(course['S']) if pd.notna(course['S']) else 0    
    c = int(course['C']) if pd.notna(course['C']) else 0    
    if s > 0 and l == 0 and t == 0 and p == 0:
        return 0, 0, 0, 0
    lecture_sessions = 0
    if l > 0:
        lecture_sessions = max(1, round(l * 2/3))  
    tutorial_sessions = t  
    lab_sessions = p // 2  
    self_study_sessions = s // 4 if (l > 0 or t > 0 or p > 0) else 0
    
    return lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions

def select_faculty(faculty_str):
    
    if '/' in faculty_str:
        faculty_options = [f.strip() for f in faculty_str.split('/')]
        return faculty_options[0]  
    return faculty_str

class UnscheduledComponent:
    def __init__(self, department, semester, code, name, faculty, component_type, sessions, section='', reason=''):
        self.department = department
        self.semester = semester
        self.code = code
        self.name = name
        self.faculty = faculty 
        self.component_type = component_type
        self.sessions = sessions
        self.section = section
        self.reason = reason
        
    def __eq__(self, other):
        if not isinstance(other, UnscheduledComponent):
            return False
        return (self.department == other.department and
                self.semester == other.semester and
                self.code == other.code and
                self.component_type == other.component_type and
                self.section == other.section)
    
    def __hash__(self):
        return hash((self.department, self.semester, self.code, self.component_type, self.section))

def unscheduled_reason(course, department, semester, professor_schedule, rooms, component_type, check_attempts):
    
    faculty = course['Faculty']
    code = str(course['Course Code'])
    faculty_slots_used = 0
    for day in range(len(DAYS)):
        if faculty in professor_schedule and day in professor_schedule[faculty]:
            faculty_slots_used += len(professor_schedule[faculty][day])
    if faculty_slots_used > 20:  
        return f"Faculty '{faculty}' already has {faculty_slots_used/2:.1f} hours of teaching scheduled"
    if component_type == 'LAB':
        lab_rooms_available = False
        for _, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                lab_rooms_available = True
                break
        
        if not lab_rooms_available:
            return "No suitable lab rooms available in the system"
        lab_rooms_free_slots = 0
        for rid, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                total_slots = len(DAYS) * (len(TIME_SLOTS) - LAB_DURATION)
                used_slots = sum(len(room['schedule'].get(day, [])) for day in range(len(DAYS)))
                lab_rooms_free_slots += (total_slots - used_slots)
        
        if lab_rooms_free_slots < 5:  
            return f"Lab rooms almost fully booked ({lab_rooms_free_slots} slots left)"
    if 'total_students' in course and pd.notna(course['total_students']):
        try:
            total_students = int(course['total_students'])
            if total_students > 100:
                large_rooms_available = False
                for _, room in rooms.items():
                    if room['type'].upper() == 'SEATER_120' or room['type'].upper() == 'SEATER_240':
                        large_rooms_available = True
                        break
                
                if not large_rooms_available:
                    return f"No rooms available with capacity for {total_students} students"
        except (ValueError, TypeError):
            pass
    if check_attempts > 800:  
        return f"No suitable timeslot found after {check_attempts} attempts - heavy scheduling conflicts"
    duration_map = {
        'LEC': f"{LECTURE_DURATION/2} hour",
        'LAB': f"{LAB_DURATION/2} hour",
        'TUT': f"{TUTORIAL_DURATION/2} hour"
    }
    duration_str = duration_map.get(component_type, "")
    
    return f"Could not find compatible {duration_str} timeslot for {code} {component_type} with faculty {faculty}"


def arrange_unscheduled_components(unscheduled_components, professor_schedule, timetable, rooms, batch_info,
                                  department=None, semester=None, section=None, elective_schedule_sync=None, num_sections=1):
    """
    Try a relaxed scheduling pass for remaining unscheduled components.
    This pass relaxes some constraints (skips faculty-gap and daily component caps)
    but still avoids double-booking faculty or rooms and respects break times.

    Parameters:
    - unscheduled_components: set of UnscheduledComponent
    - professor_schedule: dict mapping faculty -> {day: set(slots)}
    - timetable: current timetable data structure for this section
    - rooms: rooms dict from load_rooms()
    - batch_info: batch info mapping
    - department/semester/section: optional filters to only try components for this section
    """
    if not unscheduled_components or rooms is None:
        return

    scheduled_count = 0
    for comp in list(unscheduled_components):
        try:
            if department is not None and comp.department != department:
                continue
            if semester is not None and comp.semester != semester:
                continue
            if section is not None and comp.section != section:
                continue
        except Exception:
            pass

        duration = {
            'LEC': LECTURE_DURATION,
            'TUT': TUTORIAL_DURATION,
            'LAB': LAB_DURATION,
            'SS': SELF_STUDY_DURATION
        }.get(comp.component_type, LECTURE_DURATION)
        morning_slot_start = 0  # 9:00 AM corresponds to slot 0
        morning_slot_duration = 3  # 9:00-10:30 is 3 slots (90 minutes)
        candidate_slots = []
        if duration <= morning_slot_duration:
            candidate_slots.append(morning_slot_start)
        other_slots = get_candidate_start_slots(duration, True)
        if other_slots:
            candidate_slots.extend([s for s in other_slots if s != morning_slot_start])
        if not candidate_slots:
            max_start = len(TIME_SLOTS) - duration
            if max_start >= 0:
                candidate_slots = list(range(max_start + 1))

        scheduled_flag = False
        faculty = comp.faculty
        if faculty not in professor_schedule:
            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}

        for day_idx in range(len(DAYS)):
            if scheduled_flag:
                break
            for start_slot in candidate_slots:
                if start_slot > len(TIME_SLOTS) - duration:
                    continue
                conflict = False
                for i in range(duration):
                    cs = start_slot + i
                    if cs in professor_schedule[faculty][day_idx]:
                        conflict = True
                        break
                    if timetable[day_idx][cs]['type'] is not None:
                        conflict = True
                        break
                    if is_break_time(TIME_SLOTS[cs], comp.semester):
                        conflict = True
                        break
                if conflict:
                    continue

                room_type = 'LECTURE_ROOM' if comp.component_type in ['LEC', 'TUT', 'SS'] else get_required_room_type({'P': 1})
                room_id = find_suitable_room(room_type, comp.department, comp.semester,
                                             day_idx, start_slot, duration,
                                             rooms, batch_info, timetable, comp.code)
                if not room_id:
                    continue
                for i in range(duration):
                    professor_schedule[faculty][day_idx].add(start_slot + i)
                    timetable[day_idx][start_slot + i]['type'] = comp.component_type
                    timetable[day_idx][start_slot + i]['code'] = comp.code if i == 0 else ''
                    timetable[day_idx][start_slot + i]['name'] = comp.name if i == 0 else ''
                    timetable[day_idx][start_slot + i]['faculty'] = faculty if i == 0 else ''
                    timetable[day_idx][start_slot + i]['classroom'] = room_id if i == 0 else ''
                if comp.code.startswith('B0'):
                    print(f"DEBUG: B0 course {comp.code} section {comp.section} SCHEDULED in relaxed pass at Day {day_idx}, Slot {start_slot}, Room {room_id}")
                if elective_schedule_sync is not None:
                    is_basket = is_basket_course(comp.code)
                    is_elective = is_elective_or_minor_course(comp.code, comp.name, comp.department)
                    if is_basket or is_elective:
                        sync_key = (comp.semester, comp.code)  # Semester-level sync
                    else:
                        sync_key = (comp.department, comp.semester, comp.code)  # Department-level sync
                    
                    if (is_basket or is_elective) and sync_key not in elective_schedule_sync:
                        elective_schedule_sync[sync_key] = (day_idx, start_slot)
                        if comp.code.startswith('B0'):
                            print(f"DEBUG: B0 course {comp.code} - Registered sync schedule in relaxed pass for semester {comp.semester}: Day {day_idx}, Slot {start_slot}")

                try:
                    unscheduled_components.remove(comp)
                except KeyError:
                    pass
                scheduled_count += 1
                scheduled_flag = True
                break

    if scheduled_count:
        print(f"Arranged {scheduled_count} previously unscheduled component(s) in a relaxed pass")

if __name__ == "__main__":
    generate_all_timetables()
    check_unscheduled_courses()
    generate_faculty_timetables()