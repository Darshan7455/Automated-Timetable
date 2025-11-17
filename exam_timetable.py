import pandas as pd
import sys
import json
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from collections import defaultdict
import csv
import io
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
EXAM_DURATION_HOURS = 3  # Default 3 hours per exam
EXAM_SLOTS_PER_DAY = 2  # Morning and afternoon slots
MORNING_SLOT_START = "09:00"
AFTERNOON_SLOT_START = "14:00"

def load_config():
    """Load exam configuration from exam_config.json if it exists"""
    config_path = 'exam_config.json'
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r') as f:
                config = json.load(f)
                return config.get('exam_duration_minutes', 180) // 60, config.get('exam_slots_per_day', 2)
        except:
            pass
    return EXAM_DURATION_HOURS, EXAM_SLOTS_PER_DAY

def parse_dates(date_input):
    """Parse exam dates from various input formats"""
    dates = []
    date_formats = [
        '%Y-%m-%d',      # YYYY-MM-DD
        '%d-%m-%Y',      # DD-MM-YYYY
        '%Y/%m/%d',      # YYYY/MM/DD
        '%d/%m/%Y',      # DD/MM/YYYY
        '%d-%m-%y',      # DD-MM-YY (assumes 20XX)
        '%Y-%m-%d',      # YYYY-MM-DD (alternative)
    ]
    
    def try_parse_date(date_str):
        """Try to parse a date string with multiple formats"""
        for fmt in date_formats:
            try:
                parsed_datetime = datetime.strptime(date_str, fmt)
                if parsed_datetime.year < 100:
                    parsed_datetime = parsed_datetime.replace(year=2000 + parsed_datetime.year)
                return parsed_datetime.date()
            except:
                continue
        return None
    if isinstance(date_input, str) and date_input.startswith('['):
        try:
            date_list = json.loads(date_input)
            for date_str in date_list:
                parsed = try_parse_date(date_str)
                if parsed:
                    dates.append(parsed)
            return sorted(dates) if dates else []
        except:
            pass
    if isinstance(date_input, str) and ',' in date_input:
        date_strings = [d.strip() for d in date_input.split(',')]
        for date_str in date_strings:
            parsed = try_parse_date(date_str)
            if parsed:
                dates.append(parsed)
        return sorted(dates) if dates else []
    if isinstance(date_input, str):
        parsed = try_parse_date(date_input)
        if parsed:
            dates.append(parsed)
        return dates
    if isinstance(date_input, list):
        for date_item in date_input:
            if isinstance(date_item, str):
                parsed = try_parse_date(date_item)
                if parsed:
                    dates.append(parsed)
            elif hasattr(date_item, 'date'):
                dates.append(date_item.date())
        return sorted(dates) if dates else []
    
    return dates

def get_exam_dates():
    """Get exam dates from command line arguments or config file"""
    exam_dates = []
    if len(sys.argv) > 1:
        date_input = sys.argv[1]
        print(f"DEBUG: Reading dates from command line: {date_input}")
        exam_dates = parse_dates(date_input)
    if not exam_dates:
        if os.path.exists('exam_dates.json'):
            try:
                print("DEBUG: Reading dates from exam_dates.json")
                with open('exam_dates.json', 'r') as f:
                    data = json.load(f)
                    print(f"DEBUG: JSON data: {data}")
                    exam_dates = parse_dates(data.get('dates', []))
                    print(f"DEBUG: Parsed dates: {exam_dates}")
            except Exception as e:
                print(f"DEBUG: Error reading exam_dates.json: {e}")
                pass
    if not exam_dates:
        print("DEBUG: No dates provided, using defaults (next 10 weekdays)")
        today = datetime.today().date()
        current_date = today
        while len(exam_dates) < 10:
            if current_date.weekday() < 5:  # Monday to Friday
                exam_dates.append(current_date)
            current_date += timedelta(days=1)
    
    return sorted(exam_dates)

def load_courses():
    """Load courses from combined.csv"""
    try:
        if not os.path.exists('combined.csv'):
            raise FileNotFoundError("combined.csv not found in current directory")
        
        df = pd.read_csv('combined.csv')
        required_columns = ['Course Code', 'Course Name', 'Faculty', 'Department', 'Semester']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns in combined.csv: {', '.join(missing_columns)}")
        if 'Schedule' in df.columns:
            df = df[df['Schedule'].fillna('Yes').str.upper() == 'YES']
        else:
            print("Warning: 'Schedule' column not found, including all courses")
        
        if df.empty:
            raise ValueError("No courses found in combined.csv after filtering")
        
        return df
    except FileNotFoundError as e:
        print(f"Error: {e}")
        raise
    except Exception as e:
        print(f"Error loading courses: {e}")
        raise

def load_rooms():
    """Load rooms from rooms.csv"""
    rooms = {}
    try:
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['type'].upper() != 'LIBRARY':  # Exclude library
                    rooms[row['id']] = {
                        'roomNumber': row['roomNumber'],
                        'capacity': int(row['capacity']),
                        'type': row['type']
                    }
    except FileNotFoundError:
        print("Warning: rooms.csv not found, using default room allocation")
    except Exception as e:
        print(f"Warning: Error loading rooms: {e}")
    
    return rooms

def is_elective_course(course_code, department=None):
    """Determine if a course should be treated as an elective"""
    if not course_code:
        return False
    code = str(course_code).strip().upper()
    dept = str(department).strip().upper() if department is not None else ""
    
    if dept == "ELECTIVE":
        return True
    if code.startswith('B') and '-' in code:
        return True
    if 'ELECTIVE' in code:
        return True
    return False

def calculate_required_rooms_with_sharing(course, rooms, room_capacity_info):
    """Calculate room allocation allowing partial room usage (room sharing between courses)
    room_capacity_info: {room_id: {'remaining': int, 'courses': [(code, students)]}}
    Returns list of (room_id, student_count) tuples
    """
    try:
        total_students = int(course['total_students']) if pd.notna(course['total_students']) else 70
    except:
        total_students = 70
    
    if not rooms or not room_capacity_info:
        return []
    
    required_rooms = []
    remaining_students = total_students
    available_rooms = []
    excluded_rooms = ['C002', 'C003', 'C004']
    
    for room_id, usage_info in room_capacity_info.items():
        if room_id in rooms and usage_info['remaining'] > 0:
            room_info = rooms[room_id]
            room_type = room_info['type'].upper()
            room_number = room_info.get('roomNumber', '')
            
            if room_type == 'LECTURE_ROOM' and room_number not in excluded_rooms:
                room_capacity = room_info['capacity']
                utilization = room_capacity - usage_info['remaining']
                available_rooms.append((room_id, usage_info['remaining'], utilization, room_info))
    
    if not available_rooms:
        return []
    sorted_rooms = sorted(available_rooms, key=lambda x: (-x[2], -x[1]))
    for room_id, remaining_capacity, utilization, room_info in sorted_rooms:
        if remaining_students <= 0:
            break
        students_to_allocate = min(remaining_students, remaining_capacity)
        required_rooms.append((room_id, students_to_allocate))
        remaining_students -= students_to_allocate
    
    return required_rooms

def calculate_required_rooms(course, rooms, available_rooms_filter=None):
    """Calculate how many rooms are needed for a course based on rooms.csv data
    If available_rooms_filter is provided, only use rooms from that set (for finding available rooms)
    """
    try:
        total_students = int(course['total_students']) if pd.notna(course['total_students']) else 70
    except:
        total_students = 70
    
    if not rooms:
        return []
    required_rooms = []
    remaining_students = total_students
    available_rooms = []
    excluded_rooms = ['C002', 'C003', 'C004']  # Large halls to exclude
    
    for room_id, room_info in rooms.items():
        room_type = room_info['type'].upper()
        room_number = room_info.get('roomNumber', '')
        if room_type == 'LECTURE_ROOM' and room_number not in excluded_rooms:
            if available_rooms_filter is None or room_id in available_rooms_filter:
                available_rooms.append((room_id, room_info))
    
    if not available_rooms:
        return []
    sorted_rooms = sorted(available_rooms, key=lambda x: x[1]['capacity'], reverse=True)
    used_room_ids = set()
    for room_id, room_info in sorted_rooms:
        room_capacity = room_info['capacity']
        if room_capacity >= total_students:
            required_rooms.append((room_id, total_students))
            return required_rooms
    for room_id, room_info in sorted_rooms:
        if remaining_students <= 0:
            break
        
        if room_id in used_room_ids:
            continue
        
        room_capacity = room_info['capacity']
        if room_capacity >= remaining_students:
            required_rooms.append((room_id, remaining_students))
            remaining_students = 0
            used_room_ids.add(room_id)
            break
        else:
            required_rooms.append((room_id, room_capacity))
            remaining_students -= room_capacity
            used_room_ids.add(room_id)
    if remaining_students > 0:
        for room_id, room_info in sorted_rooms:
            if remaining_students <= 0:
                break
            if room_id not in used_room_ids:
                room_capacity = room_info['capacity']
                students_to_assign = min(remaining_students, room_capacity)
                required_rooms.append((room_id, students_to_assign))
                remaining_students -= students_to_assign
                used_room_ids.add(room_id)
    return required_rooms

def check_conflicts(exam_schedule, course, date, slot, rooms_needed):
    """Check if scheduling this exam would cause conflicts
    NOTE: With room sharing, room conflicts are NOT checked here - 
    we check capacity availability separately"""
    course_code = str(course['Course Code'])
    faculty = str(course['Faculty'])
    department = str(course['Department'])
    semester = str(course['Semester'])
    for existing_exam in exam_schedule:
        if (existing_exam['date'] == date and 
            existing_exam['slot'] == slot and
            existing_exam['faculty'] == faculty):
            return True, "Faculty conflict"
    is_elective = is_elective_course(course_code, department)
    
    if not is_elective:
        for existing_exam in exam_schedule:
            existing_is_elective = is_elective_course(existing_exam['course_code'], existing_exam['department'])
            if not existing_is_elective:
                if (existing_exam['date'] == date and 
                    existing_exam['slot'] == slot and
                    existing_exam['department'] == department and
                    existing_exam['semester'] == semester):
                    return True, "Student conflict"
    
    return False, None

def determine_unscheduled_reason(course, exam_dates, time_slots, exam_schedule, rooms, rooms_needed):
    """Determine the specific reason why an exam couldn't be scheduled"""
    course_code = str(course['Course Code'])
    faculty = str(course['Faculty'])
    department = str(course['Department'])
    semester = str(course['Semester'])
    if not rooms:
        return "No rooms available in rooms.csv"
    
    if not rooms_needed:
        return "Could not calculate required rooms from rooms.csv data"
    required_room_ids = [room_id for room_id, _ in rooms_needed]
    missing_rooms = [room_id for room_id in required_room_ids if room_id not in rooms]
    if missing_rooms:
        return f"Required rooms not found in rooms.csv: {', '.join(missing_rooms)}"
    available_room_capacity = 0
    excluded_rooms = ['C002', 'C003', 'C004']
    
    for room_id, room_info in rooms.items():
        room_type = room_info['type'].upper()
        room_number = room_info.get('roomNumber', '')
        if room_type == 'LECTURE_ROOM' and room_number not in excluded_rooms:
            available_room_capacity += room_info['capacity']
    
    try:
        total_students = int(course['total_students']) if pd.notna(course['total_students']) else 70
    except:
        total_students = 70
    
    if available_room_capacity < total_students:
        return f"Insufficient room capacity. Need {total_students} seats, only {available_room_capacity} available"
    faculty_conflicts = 0
    for date in exam_dates:
        for slot in time_slots:
            for existing_exam in exam_schedule:
                if (existing_exam['date'] == date and 
                    existing_exam['slot'] == slot and
                    existing_exam['faculty'] == faculty):
                    faculty_conflicts += 1
    
    if faculty_conflicts >= len(exam_dates) * len(time_slots):
        return f"Faculty '{faculty}' has conflicts in all available time slots"
    student_conflicts = 0
    for date in exam_dates:
        for slot in time_slots:
            for existing_exam in exam_schedule:
                if (existing_exam['date'] == date and 
                    existing_exam['slot'] == slot and
                    existing_exam['department'] == department and
                    existing_exam['semester'] == semester):
                    student_conflicts += 1
    
    if student_conflicts >= len(exam_dates) * len(time_slots):
        return f"Department {department} Semester {semester} has conflicts in all available time slots"
    required_room_ids = [room_id for room_id, _ in rooms_needed]
    room_conflicts = 0
    for date in exam_dates:
        for slot in time_slots:
            for existing_exam in exam_schedule:
                if (existing_exam['date'] == date and 
                    existing_exam['slot'] == slot):
                    existing_rooms = existing_exam.get('rooms', [])
                    if any(room_id in existing_rooms for room_id in required_room_ids):
                        room_conflicts += 1
    
    if room_conflicts >= len(exam_dates) * len(time_slots):
        return f"Required rooms are occupied in all available time slots"
    return "No available slot or room combination found"

def generate_exam_timetable(exam_dates, exam_duration_hours, slots_per_day):
    """Generate exam timetable"""
    courses_df = load_courses()
    rooms = load_rooms()
    
    exam_schedule = []
    unscheduled_exams = []
    time_slots = []
    if slots_per_day >= 1:
        time_slots.append(MORNING_SLOT_START)
    if slots_per_day >= 2:
        time_slots.append(AFTERNOON_SLOT_START)
    courses_df['total_students_num'] = courses_df['total_students'].fillna(70).astype(int)
    courses_df['is_elective'] = courses_df.apply(
        lambda row: is_elective_course(row['Course Code'], row['Department']), axis=1
    )
    
    core_courses = courses_df[courses_df['is_elective'] == False].copy()
    elective_courses = courses_df[courses_df['is_elective'] == True].copy()
    
    core_courses = core_courses.sort_values(['total_students_num', 'Department', 'Semester'], ascending=[False, True, True])
    elective_courses = elective_courses.sort_values(['total_students_num', 'Department', 'Semester'], ascending=[False, True, True])
    room_usage = {}  # {date: {slot: {room_id: {'remaining': capacity, 'courses': [(course_code, students)]}}}}
    for date in exam_dates:
        room_usage[date] = {}
        for slot in time_slots:
            room_usage[date][slot] = {}
            for room_id, room_info in rooms.items():
                room_type = room_info['type'].upper()
                room_number = room_info.get('roomNumber', '')
                excluded_rooms = ['C002', 'C003', 'C004']
                if room_type == 'LECTURE_ROOM' and room_number not in excluded_rooms:
                    room_usage[date][slot][room_id] = {'remaining': room_info['capacity'], 'courses': []}
    max_passes = 5  # Try up to 5 passes to schedule remaining exams
    remaining_courses = core_courses.copy()
    
    print(f"Attempting to schedule {len(remaining_courses)} core courses and {len(elective_courses)} elective courses...")
    
    for pass_num in range(max_passes):
        if remaining_courses.empty:
            break
        
        print(f"Pass {pass_num + 1}: Attempting to schedule {len(remaining_courses)} remaining courses...")
        courses_to_remove = []
        for _, course in remaining_courses.iterrows():
            course_code = str(course['Course Code'])
            course_name = str(course['Course Name'])
            faculty = str(course['Faculty'])
            department = str(course['Department'])
            semester = str(course['Semester'])
            
            scheduled = False
            for date in exam_dates:
                if scheduled:
                    break
                for slot in time_slots:
                    if scheduled:
                        break
                    available_room_ids = set()
                    for room_id, usage_info in room_usage[date][slot].items():
                        if usage_info['remaining'] > 0:  # Room has some capacity left
                            available_room_ids.add(room_id)
                    rooms_needed = calculate_required_rooms_with_sharing(course, rooms, room_usage[date][slot])
                    
                    if not rooms_needed:
                        continue
                    has_conflict, conflict_reason = check_conflicts(exam_schedule, course, date, slot, rooms_needed)
                    
                    if not has_conflict:
                        required_room_ids = [room_id for room_id, _ in rooms_needed]
                        rooms_available = True
                        
                        for room_id, student_count in rooms_needed:
                            if room_id not in rooms:
                                rooms_available = False
                                break
                            if room_id not in room_usage[date][slot]:
                                rooms_available = False
                                break
                            if room_usage[date][slot][room_id]['remaining'] < student_count:
                                rooms_available = False
                                break
                        
                        if rooms_available:
                            exam_entry = {
                                'date': date,
                                'slot': slot,
                                'course_code': course_code,
                                'course_name': course_name,
                                'faculty': faculty,
                                'department': department,
                                'semester': semester,
                                'rooms': required_room_ids,
                                'rooms_detail': rooms_needed,
                                'total_students': int(course['total_students']) if pd.notna(course['total_students']) else 70
                            }
                            exam_schedule.append(exam_entry)
                            for room_id, student_count in rooms_needed:
                                room_usage[date][slot][room_id]['remaining'] -= student_count
                                room_usage[date][slot][room_id]['courses'].append((course_code, student_count))
                            
                            scheduled = True
                            courses_to_remove.append(course_code)
            
            if not scheduled and pass_num == max_passes - 1:
                for date in exam_dates:
                    if scheduled:
                        break
                    for slot in time_slots:
                        if scheduled:
                            break
                        has_faculty_conflict = False
                        has_student_conflict = False
                        
                        for existing_exam in exam_schedule:
                            if (existing_exam['date'] == date and 
                                existing_exam['slot'] == slot and
                                existing_exam['faculty'] == faculty):
                                has_faculty_conflict = True
                                break
                            if (existing_exam['date'] == date and 
                                existing_exam['slot'] == slot and
                                existing_exam['department'] == department and
                                existing_exam['semester'] == semester):
                                has_student_conflict = True
                                break
                        
                        if not has_faculty_conflict and not has_student_conflict:
                            rooms_allocation = calculate_required_rooms_with_sharing(course, rooms, room_usage[date][slot])
                            
                            if rooms_allocation:  # If we can allocate with remaining capacity
                                required_room_ids = [room_id for room_id, _ in rooms_allocation]
                                can_allocate = True
                                for room_id, student_count in rooms_allocation:
                                    if room_usage[date][slot][room_id]['remaining'] < student_count:
                                        can_allocate = False
                                        break
                                
                                if can_allocate:
                                    exam_entry = {
                                        'date': date,
                                        'slot': slot,
                                        'course_code': course_code,
                                        'course_name': course_name,
                                        'faculty': faculty,
                                        'department': department,
                                        'semester': semester,
                                        'rooms': required_room_ids,
                                        'rooms_detail': rooms_allocation,
                                        'total_students': total_students
                                    }
                                    exam_schedule.append(exam_entry)
                                    for room_id, student_count in rooms_allocation:
                                        room_usage[date][slot][room_id]['remaining'] -= student_count
                                        room_usage[date][slot][room_id]['courses'].append((course_code, student_count))
                                    
                                    scheduled = True
                                    courses_to_remove.append(course_code)
                                    break
        if courses_to_remove:
            print(f"  Scheduled {len(courses_to_remove)} courses in pass {pass_num + 1}")
            remaining_courses = remaining_courses[~remaining_courses['Course Code'].astype(str).isin(courses_to_remove)]
        else:
            print(f"  No courses scheduled in pass {pass_num + 1}")
    
    print(f"\nFinal scheduling result: {len(exam_schedule)} scheduled, {len(remaining_courses)} remaining")
    for _, course in remaining_courses.iterrows():
        course_code = str(course['Course Code'])
        course_name = str(course['Course Name'])
        faculty = str(course['Faculty'])
        department = str(course['Department'])
        semester = str(course['Semester'])
        
        rooms_needed = calculate_required_rooms(course, rooms)
        reason = determine_unscheduled_reason(course, exam_dates, time_slots, exam_schedule, rooms, rooms_needed)
        unscheduled_exams.append({
            'course_code': course_code,
            'course_name': course_name,
            'faculty': faculty,
            'department': department,
            'semester': semester,
            'total_students': int(course['total_students']) if pd.notna(course['total_students']) else 70,
            'rooms_required': len(rooms_needed),
            'reason': reason
        })
    elective_unscheduled = schedule_elective_exams(
        elective_courses,
        exam_dates,
        time_slots,
        rooms,
        exam_schedule,
        room_usage
    )
    unscheduled_exams.extend(elective_unscheduled)
    
    return exam_schedule, unscheduled_exams

def schedule_elective_exams(elective_courses, exam_dates, time_slots, rooms, exam_schedule, room_usage):
    """Schedule elective courses to ensure they occupy different day/time slots when possible"""
    unscheduled = []
    
    if elective_courses.empty:
        return unscheduled
    
    print(f"\nScheduling {len(elective_courses)} elective courses...")
    electives_sorted = elective_courses.sort_values(['total_students_num', 'Course Code'], ascending=[False, True])
    
    elective_day_counts = {date: 0 for date in exam_dates}
    
    for _, course in electives_sorted.iterrows():
        course_code = str(course['Course Code'])
        course_name = str(course['Course Name'])
        faculty = str(course['Faculty'])
        department = str(course['Department'])
        semester = str(course['Semester'])
        try:
            total_students = int(course['total_students']) if pd.notna(course['total_students']) else 70
        except:
            total_students = 70
        
        scheduled = False
        slot_candidates = []
        for date in exam_dates:
            for slot in time_slots:
                total_exams_in_slot = sum(1 for exam in exam_schedule if exam['date'] == date and exam['slot'] == slot)
                elective_exams_in_slot = sum(1 for exam in exam_schedule if exam['date'] == date and exam['slot'] == slot and is_elective_course(exam['course_code']))
                day_elective_count = elective_day_counts.get(date, 0)
                candidate = (day_elective_count, elective_exams_in_slot, total_exams_in_slot, date, slot)
                slot_candidates.append(candidate)
        
        slot_candidates.sort()
        
        for day_elective_count, elective_count, total_count, date, slot in slot_candidates:
            rooms_allocation = calculate_required_rooms_with_sharing(course, rooms, room_usage[date][slot])
            if not rooms_allocation:
                continue
            
            conflict, _ = check_conflicts(exam_schedule, course, date, slot, rooms_allocation)
            if conflict:
                continue
            can_allocate = True
            for room_id, student_count in rooms_allocation:
                if room_usage[date][slot][room_id]['remaining'] < student_count:
                    can_allocate = False
                    break
            
            if not can_allocate:
                continue
            
            required_room_ids = [room_id for room_id, _ in rooms_allocation]
            exam_entry = {
                'date': date,
                'slot': slot,
                'course_code': course_code,
                'course_name': course_name,
                'faculty': faculty,
                'department': department,
                'semester': semester,
                'rooms': required_room_ids,
                'rooms_detail': rooms_allocation,
                'total_students': total_students
            }
            exam_schedule.append(exam_entry)
            for room_id, student_count in rooms_allocation:
                room_usage[date][slot][room_id]['remaining'] -= student_count
                room_usage[date][slot][room_id]['courses'].append((course_code, student_count))
            
            scheduled = True
            elective_day_counts[date] = elective_day_counts.get(date, 0) + 1
            break
        
        if not scheduled:
            unscheduled.append({
                'course_code': course_code,
                'course_name': course_name,
                'faculty': faculty,
                'department': department,
                'semester': semester,
                'total_students': total_students,
                'rooms_required': 0,
                'reason': 'Could not find suitable slot for elective course'
            })
    
    if unscheduled:
        print(f"Electives unscheduled after dedicated pass: {len(unscheduled)}")
    else:
        print("All elective courses scheduled successfully.")
    
    return unscheduled

def generate_seat_arrangement(exam, rooms):
    """Generate detailed seat arrangement for an exam
    With room sharing, this shows which students from THIS course are in each room
    """
    seat_arrangement = []
    rooms_detail = exam.get('rooms_detail', [])
    total_students = exam.get('total_students', 0)
    
    if not rooms_detail:
        return seat_arrangement
    department = exam.get('department', 'XXX')
    semester = exam.get('semester', '0')
    course_code = exam.get('course_code', 'XXX')
    
    current_student = 1
    
    for room_id, student_count in rooms_detail:
        if room_id not in rooms:
            continue
            
        room_info = rooms[room_id]
        room_number = room_info['roomNumber']
        room_capacity = room_info['capacity']
        seats_per_row = 10  # Default assumption
        seats = []
        for i in range(student_count):
            seat_number = i + 1
            student_id = f"{department}-S{semester}-{current_student:03d}"
            seats.append({
                'student_id': student_id,
                'seat_number': seat_number,
                'row': (seat_number - 1) // seats_per_row + 1,
                'column': (seat_number - 1) % seats_per_row + 1
            })
            current_student += 1
        
        seat_arrangement.append({
            'room_id': room_id,
            'room_number': room_number,
            'room_capacity': room_capacity,
            'students_allocated': student_count,
            'seats': seats,
            'roll_range': f"{seats[0]['student_id']} to {seats[-1]['student_id']}" if seats else "N/A",
            'course_code': course_code,  # Add course code for room sharing display
            'is_shared': False  # Will be updated if room has multiple courses
        })
    
    return seat_arrangement

def create_excel_timetable(exam_schedule, unscheduled_exams, exam_dates, exam_duration_hours, slots_per_day, rooms):
    """Create Excel file for exam timetable"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
    except Exception as e:
        print(f"Error creating workbook: {e}")
        raise
    overview_sheet = wb.create_sheet(title="Overview")
    overview_sheet.append(["Exam Timetable Overview"])
    overview_sheet.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview_sheet.append([])
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    date_range_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    title_cell = overview_sheet['A1']
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    if exam_dates:
        sorted_dates = sorted(exam_dates)
        start_date = sorted_dates[0]
        end_date = sorted_dates[-1]
        total_days = len(exam_dates)
        unique_days = len(set(exam_dates))
        
        row_idx = 4
        overview_sheet.append([])
        overview_sheet.append(["📅 EXAM DATE RANGE"])
        date_range_title_row = overview_sheet.max_row
        overview_sheet.cell(row=date_range_title_row, column=1).font = header_font
        overview_sheet.cell(row=date_range_title_row, column=1).fill = date_range_fill
        
        overview_sheet.append(["  Start Date:", start_date.strftime('%Y-%m-%d (%A)')])
        overview_sheet.append(["  End Date:", end_date.strftime('%Y-%m-%d (%A)')])
        overview_sheet.append([])
    
    overview_sheet.append(["📊 EXAM STATISTICS"])
    stats_title_row = overview_sheet.max_row
    overview_sheet.cell(row=stats_title_row, column=1).font = header_font
    overview_sheet.cell(row=stats_title_row, column=1).fill = date_range_fill
    
    overview_sheet.append(["  Total Exams Scheduled:", len(exam_schedule)])
    overview_sheet.append(["  Total Unscheduled Exams:", len(unscheduled_exams)])
    overview_sheet.append([])
    overview_sheet.append(["📋 ALL EXAM DATES:"])
    dates_title_row = overview_sheet.max_row
    overview_sheet.cell(row=dates_title_row, column=1).font = header_font
    overview_sheet.cell(row=dates_title_row, column=1).fill = date_range_fill
    
    for date in sorted(exam_dates):
        day_name = date.strftime('%A')
        overview_sheet.append([f"    • {date.strftime('%Y-%m-%d')} ({day_name})"])
    sharing_sheet = wb.create_sheet(title="Room Sharing Info")
    sharing_sheet.append(["Room Sharing Summary"])
    sharing_sheet.append(["Shows which rooms have multiple courses during the same time slot"])
    sharing_sheet.append([])
    room_sharing_data = {}  # {(date, slot, room_id): [(course_code, students)]}
    
    for exam in exam_schedule:
        date = exam['date']
        slot = exam['slot']
        course_code = exam['course_code']
        rooms_detail = exam.get('rooms_detail', [])
        
        for room_id, student_count in rooms_detail:
            key = (date, slot, room_id)
            if key not in room_sharing_data:
                room_sharing_data[key] = []
            room_sharing_data[key].append((course_code, student_count))
    shared_rooms = {k: v for k, v in room_sharing_data.items() if len(v) > 1}
    
    if shared_rooms:
        sharing_sheet.append([f"Total Shared Room Instances: {len(shared_rooms)}"])
        sharing_sheet.append([])
        sharing_header = ['Date', 'Day', 'Time Slot', 'Room', 'Courses in Room (Students)']
        sharing_sheet.append(sharing_header)
        
        header_row = sharing_sheet.max_row
        for col in range(1, 6):
            cell = sharing_sheet.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for (date, slot, room_id), courses in sorted(shared_rooms.items()):
            room_number = rooms.get(room_id, {}).get('roomNumber', room_id)
            room_capacity = rooms.get(room_id, {}).get('capacity', 'N/A')
            date_str = date.strftime('%Y-%m-%d')
            day_name = date.strftime('%A')
            courses_info = ' + '.join([f"{code}({students})" for code, students in courses])
            total_students = sum(students for _, students in courses)
            courses_info += f" = {total_students}/{room_capacity} seats"
            
            row = [date_str, day_name, slot, room_number, courses_info]
            sharing_sheet.append(row)
            
            for cell in sharing_sheet[sharing_sheet.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        sharing_sheet.append([])
        sharing_sheet.append(["Benefits of Room Sharing:"])
        sharing_sheet.append(["  • Efficient use of classroom capacity"])
        sharing_sheet.append(["  • More exams can be scheduled simultaneously"])
        sharing_sheet.append(["  • Reduces scheduling conflicts"])
    else:
        sharing_sheet.append(["No rooms are currently being shared between multiple courses."])
        sharing_sheet.append(["Each room hosts only one course per time slot."])
    sharing_sheet.column_dimensions['A'].width = 12
    sharing_sheet.column_dimensions['B'].width = 12
    sharing_sheet.column_dimensions['C'].width = 12
    sharing_sheet.column_dimensions['D'].width = 12
    sharing_sheet.column_dimensions['E'].width = 60
    main_sheet = wb.create_sheet(title="Exam Schedule")
    main_sheet.append(["Indian Institute of Information Technology Dharwad"])
    main_sheet.append(["End Semester Examination - Aug-Nov 2025"])
    main_sheet.append(["AN: 03:00 PM to 04:30 PM"])
    main_sheet.append([])
    header = ['Date', 'Days', 'Course Code']
    main_sheet.append(header)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for row in range(1, 4):
        for cell in main_sheet[row]:
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in main_sheet[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    exams_by_date = {}
    for exam in exam_schedule:
        date = exam['date']
        if date not in exams_by_date:
            exams_by_date[date] = []
        exams_by_date[date].append(exam)
    sorted_dates = sorted(exams_by_date.keys())
    
    for date in sorted_dates:
        try:
            exams_for_date = exams_by_date[date]
            month_abbr = date.strftime('%b')  # Gets abbreviated month like "Nov"
            date_display = f"{date.strftime('%d')}-{month_abbr}-{date.strftime('%Y')} ({date.strftime('%A')})"
            day_name = date.strftime('%A')
            course_codes = sorted([str(exam['course_code']) for exam in exams_for_date if exam.get('course_code')])
            if not course_codes:
                continue  # Skip if no course codes
            course_codes_str = ', '.join(course_codes)
            
            row = [date_display, day_name, course_codes_str]
            main_sheet.append(row)
        except Exception as e:
            print(f"Error processing date {date}: {e}")
            import traceback
            traceback.print_exc()
            continue
        for cell in main_sheet[main_sheet.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    main_sheet.column_dimensions['A'].width = 25  # Date
    main_sheet.column_dimensions['B'].width = 15  # Days
    main_sheet.column_dimensions['C'].width = 80  # Course Code (wider for multiple codes)
    detail_sheet = wb.create_sheet(title="Detailed Schedule")
    detail_sheet.append(["Detailed Exam Schedule with Room Allocation"])
    detail_sheet.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    detail_sheet.append([])
    
    detail_header = ['Date', 'Day', 'Time Slot', 'Course Code', 'Course Name', 'Department', 'Semester', 'Faculty', 'Rooms (Seats)', 'Total Students']
    detail_sheet.append(detail_header)
    for cell in detail_sheet[detail_sheet.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    exam_schedule_sorted = sorted(exam_schedule, key=lambda x: (x['date'], x['slot']))
    
    for exam in exam_schedule_sorted:
        date_str = exam['date'].strftime('%Y-%m-%d')
        day_name = exam['date'].strftime('%A')
        rooms_list = []
        for room_id, count in exam['rooms_detail']:
            if room_id in rooms:
                room_num = rooms[room_id]['roomNumber']
                rooms_list.append(f"{room_num}({count})")
        rooms_str = ', '.join(rooms_list) if rooms_list else "N/A"
        
        row = [
            date_str,
            day_name,
            exam['slot'],
            exam['course_code'],
            exam['course_name'],
            exam['department'],
            exam['semester'],
            exam['faculty'],
            rooms_str,
            exam['total_students']
        ]
        detail_sheet.append(row)
        for cell in detail_sheet[detail_sheet.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    detail_sheet.column_dimensions['A'].width = 12
    detail_sheet.column_dimensions['B'].width = 12
    detail_sheet.column_dimensions['C'].width = 12
    detail_sheet.column_dimensions['D'].width = 15
    detail_sheet.column_dimensions['E'].width = 40
    detail_sheet.column_dimensions['F'].width = 15
    detail_sheet.column_dimensions['G'].width = 12
    detail_sheet.column_dimensions['H'].width = 25
    detail_sheet.column_dimensions['I'].width = 40
    detail_sheet.column_dimensions['J'].width = 10
    for date in exam_dates:
        date_str = date.strftime('%Y-%m-%d')
        day_name = date.strftime('%A')
        sheet_name = f"{date_str[:10]}"  # Limit to 31 chars
        if len(sheet_name) > 31:
            sheet_name = date_str[:31]
        
        date_sheet = wb.create_sheet(title=sheet_name)
        date_sheet.append([f"Exam Schedule for {date_str} ({day_name})"])
        date_sheet.append([])
        time_slots = []
        if slots_per_day >= 1:
            time_slots.append(MORNING_SLOT_START)
        if slots_per_day >= 2:
            time_slots.append(AFTERNOON_SLOT_START)
        header_row = ['Course Code', 'Course Name', 'Department', 'Semester', 'Faculty', 'Rooms', 'Students']
        date_sheet.append(header_row)
        for cell in date_sheet[date_sheet.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        date_exams = [e for e in exam_schedule if e['date'] == date]
        date_exams = sorted(date_exams, key=lambda x: x['course_code'])
        for slot in time_slots:
            slot_exams = [e for e in date_exams if e['slot'] == slot]
            
            if slot_exams:
                date_sheet.append([f"Slot: {slot}"])
                for exam in slot_exams:
                    rooms_str = ', '.join([f"{rooms[room_id]['roomNumber']} ({count})" 
                                          for room_id, count in exam['rooms_detail'] 
                                          if room_id in rooms])
                    
                    row = [
                        exam['course_code'],
                        exam['course_name'],
                        exam['department'],
                        exam['semester'],
                        exam['faculty'],
                        rooms_str,
                        exam['total_students']
                    ]
                    date_sheet.append(row)
                    for cell in date_sheet[date_sheet.max_row]:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                date_sheet.append([])  # Empty row between slots
    seat_sheet = wb.create_sheet(title="Seat Arrangement")
    seat_sheet.append(["Exam Seat Arrangement"])
    seat_sheet.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    seat_sheet.append([])
    seat_header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    room_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    for exam in exam_schedule_sorted:
        exam_header = f"{exam['course_code']} - {exam['course_name']}"
        seat_sheet.append([exam_header])
        header_row = seat_sheet.max_row
        seat_sheet.cell(row=header_row, column=1).font = Font(bold=True, size=12)
        seat_sheet.cell(row=header_row, column=1).fill = header_fill
        date_str = exam['date'].strftime('%Y-%m-%d (%A)')
        seat_sheet.append(["Date:", date_str, "Time:", exam['slot'], "Department:", exam['department'], "Semester:", exam['semester']])
        seat_sheet.append(["Faculty:", exam['faculty'], "Total Students:", exam['total_students']])
        seat_sheet.append([])
        seat_arrangement = generate_seat_arrangement(exam, rooms)
        
        if seat_arrangement:
            seat_sheet.append(["Room", "Room Number", "Capacity", "Students from This Course", "Student Roll Range"])
            header_row = seat_sheet.max_row
            for col in range(1, 6):
                cell = seat_sheet.cell(row=header_row, column=col)
                cell.fill = seat_header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for room_data in seat_arrangement:
                row = [
                    room_data['room_id'],
                    room_data['room_number'],
                    room_data['room_capacity'],
                    room_data['students_allocated'],
                    room_data['roll_range']
                ]
                seat_sheet.append(row)
                for cell in seat_sheet[seat_sheet.max_row]:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.fill = room_fill
            seat_sheet.append([])
            seat_sheet.append(["Note: Rooms may be shared with other courses during the same time slot."])
            note_row = seat_sheet.max_row
            seat_sheet.cell(row=note_row, column=1).font = Font(italic=True, size=9)
            seat_sheet.append([])
            for room_data in seat_arrangement:
                seat_sheet.append([f"Room {room_data['room_number']} - Detailed Seat Arrangement"])
                detail_header_row = seat_sheet.max_row
                seat_sheet.cell(row=detail_header_row, column=1).font = Font(bold=True, size=11)
                seat_sheet.cell(row=detail_header_row, column=1).fill = room_fill
                seat_sheet.append(["Seat Number", "Row", "Column", "Student ID"])
                detail_header_row = seat_sheet.max_row
                for col in range(1, 5):
                    cell = seat_sheet.cell(row=detail_header_row, column=col)
                    cell.fill = seat_header_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                for seat in room_data['seats']:
                    row = [
                        seat['seat_number'],
                        seat['row'],
                        seat['column'],
                        seat['student_id']
                    ]
                    seat_sheet.append(row)
                    for cell in seat_sheet[seat_sheet.max_row]:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                seat_sheet.append([])
        
        seat_sheet.append([])
        seat_sheet.append(["="*50])
        seat_sheet.append([])
    seat_sheet.column_dimensions['A'].width = 15
    seat_sheet.column_dimensions['B'].width = 15
    seat_sheet.column_dimensions['C'].width = 12
    seat_sheet.column_dimensions['D'].width = 18
    seat_sheet.column_dimensions['E'].width = 30
    seat_sheet.column_dimensions['F'].width = 15
    seat_sheet.column_dimensions['G'].width = 12
    seat_sheet.column_dimensions['H'].width = 15
    seatmap_sheet = wb.create_sheet(title="Seat Maps")
    seatmap_sheet.append(["Visual Seat Maps - Room Sharing View"])
    seatmap_sheet.append(["Shows which courses share rooms during same time slots"])
    seatmap_sheet.append([])
    slot_room_map = {}
    for exam in exam_schedule_sorted:
        date = exam['date']
        slot = exam['slot']
        for room_id, student_count in exam.get('rooms_detail', []):
            key = (date, slot, room_id)
            if key not in slot_room_map:
                slot_room_map[key] = []
            slot_room_map[key].append({
                'course_code': exam['course_code'],
                'course_name': exam['course_name'],
                'department': exam['department'],
                'semester': exam['semester'],
                'students': student_count
            })
    shared_count = 0
    single_count = 0
    
    for (date, slot, room_id), courses in sorted(slot_room_map.items()):
        room_info = rooms.get(room_id, {})
        room_number = room_info.get('roomNumber', room_id)
        room_capacity = room_info.get('capacity', 'N/A')
        
        is_shared = len(courses) > 1
        
        if is_shared:
            shared_count += 1
            date_str = date.strftime('%Y-%m-%d (%A)')
            header_text = f"🔄 SHARED: Room {room_number} | {date_str} | {slot}"
            seatmap_sheet.append([header_text])
            header_row = seatmap_sheet.max_row
            seatmap_sheet.cell(row=header_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            seatmap_sheet.cell(row=header_row, column=1).fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            seatmap_sheet.append([])
            total_students = sum(c['students'] for c in courses)
            capacity_text = f"Room Capacity: {room_capacity} | Total Students: {total_students} | Utilization: {int(total_students/room_capacity*100)}%"
            seatmap_sheet.append([capacity_text])
            cap_row = seatmap_sheet.max_row
            seatmap_sheet.cell(row=cap_row, column=1).font = Font(bold=True, size=10)
            seatmap_sheet.cell(row=cap_row, column=1).fill = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")
            seatmap_sheet.append([])
            seatmap_sheet.append(["Course Code", "Course Name", "Department", "Semester", "Students"])
            course_header_row = seatmap_sheet.max_row
            for col in range(1, 6):
                cell = seatmap_sheet.cell(row=course_header_row, column=col)
                cell.fill = seat_header_fill
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for course in courses:
                row_data = [
                    course['course_code'],
                    course['course_name'],
                    course['department'],
                    course['semester'],
                    course['students']
                ]
                seatmap_sheet.append(row_data)
                
                for cell in seatmap_sheet[seatmap_sheet.max_row]:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            seatmap_sheet.append([])
            seatmap_sheet.append(["─" * 80])
            seatmap_sheet.append([])
        else:
            single_count += 1
    seatmap_sheet.insert_rows(4)
    seatmap_sheet.cell(row=4, column=1).value = f"📊 Summary: {shared_count} shared rooms | {single_count} single-course rooms"
    seatmap_sheet.cell(row=4, column=1).font = Font(bold=True, size=11, color="0066CC")
    seatmap_sheet.append([])
    seatmap_sheet.column_dimensions['A'].width = 20
    seatmap_sheet.column_dimensions['B'].width = 40
    seatmap_sheet.column_dimensions['C'].width = 15
    seatmap_sheet.column_dimensions['D'].width = 12
    seatmap_sheet.column_dimensions['E'].width = 12
    traditional_seatmap = wb.create_sheet(title="Traditional Seat Maps")
    traditional_seatmap.append(["Traditional Visual Seat Maps by Room"])
    traditional_seatmap.append([])
    room_exam_map = {}
    for exam in exam_schedule_sorted:
        seat_arrangement = generate_seat_arrangement(exam, rooms)
        for room_data in seat_arrangement:
            room_id = room_data['room_id']
            if room_id not in room_exam_map:
                room_exam_map[room_id] = []
            room_exam_map[room_id].append({
                'exam': exam,
                'room_data': room_data
            })
    for room_id in sorted(room_exam_map.keys()):
        room_info = rooms.get(room_id, {})
        room_number = room_info.get('roomNumber', room_id)
        
        traditional_seatmap.append([f"Room {room_number} - Seat Map"])
        header_row = traditional_seatmap.max_row
        traditional_seatmap.cell(row=header_row, column=1).font = Font(bold=True, size=13)
        traditional_seatmap.cell(row=header_row, column=1).fill = header_fill
        traditional_seatmap.append([])
        for exam_data in room_exam_map[room_id]:
            exam = exam_data['exam']
            room_data = exam_data['room_data']
            
            course_info = f"{exam['course_code']} - {exam['date'].strftime('%Y-%m-%d')} {exam['slot']}"
            traditional_seatmap.append([course_info])
            info_row = traditional_seatmap.max_row
            traditional_seatmap.cell(row=info_row, column=1).font = Font(bold=True, size=11)
            seats_per_row = 10
            seats = room_data['seats']
            
            if seats:
                header = ["Row\\Col"] + [f"C{i+1}" for i in range(seats_per_row)]
                traditional_seatmap.append(header)
                header_row = traditional_seatmap.max_row
                for col in range(1, seats_per_row + 2):
                    cell = traditional_seatmap.cell(row=header_row, column=col)
                    cell.fill = seat_header_fill
                    cell.font = Font(bold=True, color="FFFFFF", size=9)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                max_row = max(seat['row'] for seat in seats)
                for row_num in range(1, max_row + 1):
                    row_data = [f"R{row_num}"]
                    for col_num in range(1, seats_per_row + 1):
                        seat = next((s for s in seats if s['row'] == row_num and s['column'] == col_num), None)
                        if seat:
                            row_data.append(seat['student_id'])
                        else:
                            row_data.append("---")
                    
                    traditional_seatmap.append(row_data)
                    for col_idx, cell in enumerate(traditional_seatmap[traditional_seatmap.max_row], 1):
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        if col_idx == 1:
                            cell.fill = room_fill
                            cell.font = Font(bold=True, size=9)
                        else:
                            cell.font = Font(size=8)
                        if col_idx > 1:
                            col_letter = get_column_letter(col_idx)
                            traditional_seatmap.column_dimensions[col_letter].width = 12
            
            traditional_seatmap.append([])
        
        traditional_seatmap.append([])
    if unscheduled_exams:
        unscheduled_sheet = wb.create_sheet(title="Unscheduled Exams")
        unscheduled_sheet.append(["Unscheduled Exams"])
        unscheduled_sheet.append(["Total Unscheduled:", len(unscheduled_exams)])
        unscheduled_sheet.append([])
        
        header = ['Course Code', 'Course Name', 'Department', 'Semester', 'Faculty', 'Total Students', 'Rooms Required', 'Reason']
        unscheduled_sheet.append(header)
        header_fill_unscheduled = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for cell in unscheduled_sheet[unscheduled_sheet.max_row]:
            cell.fill = header_fill_unscheduled
            cell.font = header_font
            cell.alignment = header_alignment
        
        for exam in unscheduled_exams:
            row = [
                exam['course_code'],
                exam['course_name'],
                exam['department'],
                exam['semester'],
                exam['faculty'],
                exam.get('total_students', 'N/A'),
                exam.get('rooms_required', 'N/A'),
                exam.get('reason', 'Unknown reason')
            ]
            unscheduled_sheet.append(row)
            for cell in unscheduled_sheet[unscheduled_sheet.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        unscheduled_sheet.column_dimensions['A'].width = 15  # Course Code
        unscheduled_sheet.column_dimensions['B'].width = 40  # Course Name
        unscheduled_sheet.column_dimensions['C'].width = 15  # Department
        unscheduled_sheet.column_dimensions['D'].width = 12  # Semester
        unscheduled_sheet.column_dimensions['E'].width = 25  # Faculty
        unscheduled_sheet.column_dimensions['F'].width = 15  # Total Students
        unscheduled_sheet.column_dimensions['G'].width = 15  # Rooms Required
        unscheduled_sheet.column_dimensions['H'].width = 50  # Reason
    for sheet in wb.worksheets:
        if sheet.title == "Overview":
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 30
        elif sheet.title == "Exam Schedule":
            sheet.column_dimensions['A'].width = 12
            sheet.column_dimensions['B'].width = 12
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 40
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 12
            sheet.column_dimensions['H'].width = 25
            sheet.column_dimensions['I'].width = 30
            sheet.column_dimensions['J'].width = 10
    for sheet in wb.worksheets:
        sheet.protection = SheetProtection(sheet=False, password=None)
    output_file = 'exam_timetable.xlsx'
    wb.save(output_file)
    print(f"\n{'='*60}")
    print(f"Exam timetable saved as {output_file}")
    print(f"{'='*60}")
    print(f"Total exams scheduled: {len(exam_schedule)}")
    print(f"Total unscheduled exams: {len(unscheduled_exams)}")
    print(f"\n📋 Excel sheets created:")
    print(f"  • Overview - Summary statistics")
    print(f"  • Exam Schedule - Date-wise exam list")
    print(f"  • Detailed Schedule - Complete exam details with rooms")
    print(f"  • Seat Arrangement - Student seat allocations")
    print(f"  • Seat Maps - Room sharing visualization (which courses share rooms)")
    print(f"  • Traditional Seat Maps - Visual grid seat maps")
    print(f"  • Date-wise sheets - Individual date schedules")
    if unscheduled_exams:
        print(f"  • Unscheduled Exams - Courses that couldn't be scheduled")
    
    if unscheduled_exams:
        print(f"\n[!] Unscheduled Exams Details:")
        for exam in unscheduled_exams:
            print(f"  - {exam['course_code']}: {exam.get('reason', 'Unknown reason')}")
    if rooms:
        total_rooms = len(rooms)
        exam_rooms = set()
        for exam in exam_schedule:
            exam_rooms.update(exam.get('rooms', []))
        used_rooms = len(exam_rooms)
        print(f"\n[*] Room Allocation Summary:")
        print(f"  - Total rooms in rooms.csv: {total_rooms}")
        print(f"  - Rooms used for exams: {used_rooms}")
        print(f"  - Available rooms: {total_rooms - used_rooms}")
    
    print(f"{'='*60}\n")
    
    return output_file

def main():
    """Main function"""
    try:
        exam_dates = get_exam_dates()
        
        if not exam_dates:
            print("Warning: No exam dates provided. Using default dates (next 10 weekdays).")
        
        print(f"Generating exam timetable for {len(exam_dates)} dates:")
        for date in exam_dates:
            print(f"  - {date.strftime('%Y-%m-%d (%A)')}")
        if not os.path.exists('combined.csv'):
            print("Error: combined.csv not found. Please upload the file first.")
            sys.exit(1)
        exam_duration_hours, slots_per_day = load_config()
        exam_schedule, unscheduled_exams = generate_exam_timetable(exam_dates, exam_duration_hours, slots_per_day)
        rooms = load_rooms()
        create_excel_timetable(exam_schedule, unscheduled_exams, exam_dates, exam_duration_hours, slots_per_day, rooms)
        
        print("\n✅ Exam timetable generated successfully!")
        print("📄 Output file: exam_timetable.xlsx")
        
    except FileNotFoundError as e:
        error_msg = f"Error: Required file not found: {e}"
        print(error_msg, file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        error_msg = f"Error generating exam timetable: {e}"
        print(error_msg, file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()

